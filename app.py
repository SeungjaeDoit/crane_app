from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response, flash, send_file, current_app
import datetime as _dt
from datetime import timedelta, datetime, date
from functools import wraps
from io import StringIO, BytesIO
from math import ceil
import json, os, uuid, csv, re, io, mimetypes, smtplib, secrets, hashlib, zipfile, urllib.request, urllib.error, unicodedata, ssl
from urllib.parse import quote as _q, urlencode
from email.message import EmailMessage
from email.header import Header
from email.utils import formataddr
from werkzeug.utils import secure_filename
from dotenv import load_dotenv, dotenv_values
from html import escape
import boto3
from botocore.config import Config 
import unicodedata
from markupsafe import escape

load_dotenv(dotenv_path=Path(__file__).with_name('.env'))
# .env 한 번만 로드
ENV_PATH = Path(__file__).resolve().parent / ".env"
load_dotenv(ENV_PATH, override=True)


mimetypes.add_type('application/x-hwp', '.hwp')
mimetypes.add_type('application/pdf', '.pdf')
mimetypes.add_type('application/vnd.ms-excel', '.xls')
mimetypes.add_type('application/vnd.openxmlformats-officedocument.wordprocessingml.document', '.docx')
mimetypes.add_type('application/vnd.openxmlformats-officedocument.presentationml.presentation', '.pptx')
mimetypes.add_type('application/vnd.hancom.hwp', '.hwp')     
mimetypes.add_type('application/vnd.hancom.hwpx', '.hwpx')   
mimetypes.add_type("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx")
mimetypes.add_type("application/vnd.ms-excel", ".xls")
mimetypes.add_type("application/haansofthwp", ".hwp")
mimetypes.add_type("application/x-hwp", ".hwp")

_ALLOWED_HANGUL = "\u3130-\u318F\uAC00-\uD7A3" 


def _safe_filename_kor(original: str) -> tuple[str, str, str]:
    """
    original -> (display_name, safe_name, ascii_fallback)
      - display_name: 메일에 보여줄 원래 이름(정규화만)
      - safe_name:    S3 키에 쓸 안전한 이름(한글/숫자/영문/.-_만 허용)
      - ascii_fallback: Content-Disposition용 ASCII 대체 이름
    """
    name = os.path.basename(original or "").strip()
    if not name:
        name = "file.bin"

    # 유니코드 정규화(한글 호환성)
    name = unicodedata.normalize("NFKC", name)

    base, ext = os.path.splitext(name)
    # 확장자는 반드시 점 포함 상태로 보존
    ext = ext if ext.startswith(".") else (("." + ext) if ext else "")

    # 한글/영문/숫자/.-_ 만 남기고 나머지는 '_' 로
    base = re.sub(fr"[^\w\.\-\s{_ALLOWED_HANGUL}]", "_", base)
    base = re.sub(r"\s+", "_", base).strip("_")
    if not base:
        base = "file"

    display_name = f"{base}{ext}"

    # S3 키에 쓸 세이프 네임(길이 제한 살짝)
    safe_name = display_name[:120]

    # ASCII fallback (메일 클라이언트 호환용)
    ascii_fallback = re.sub(r"[^\x00-\x7F]", "_", display_name)[:80]
    if not ascii_fallback:
        ascii_fallback = "file" + (ext or ".bin")

    return display_name, safe_name, ascii_fallback

def _guess_content_type(filename: str, file_obj) -> str:
    # 우선 브라우저가 준 mimetype
    if hasattr(file_obj, "mimetype") and file_obj.mimetype:
        return file_obj.mimetype
    # 파일명으로 추정
    guessed, _ = mimetypes.guess_type(filename)
    return guessed or "application/octet-stream"

def _parse_emails(to_raw: str):
    # 쉼표/공백/세미콜론 허용
    parts = [p.strip() for p in to_raw.replace(";", ",").split(",")]
    return [p for p in parts if p]

# =========================================================
# 경로/앱 설정
# =========================================================
BASE_DIR = Path(__file__).resolve().parent

# JSON 등 내부 데이터 저장소
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# 첨부 파일 보관 폴더(정적 파일)
DOCS_DIR = BASE_DIR / "static" / "docs"
DOCS_DIR.mkdir(parents=True, exist_ok=True)

DOCS_DB = BASE_DIR / "docs.json"

ALLOWED_DOC_EXTS = {
    "pdf","png","jpg","jpeg","gif","webp","txt","csv","xlsx","xls",
    "doc","docx","ppt","pptx","hwp","hwpx","zip"
}

app = Flask(__name__)

import os

def _init_config_from_env(app):
    # .env에서 읽어 app.config로 복사
    app.config.update(
        CLOUD_BACKEND=os.getenv("CLOUD_BACKEND", "").lower(),
        S3_BUCKET=os.getenv("S3_BUCKET"),
        S3_REGION=os.getenv("S3_REGION"),
        S3_EXPIRE_SECONDS=int(os.getenv("S3_EXPIRE_SECONDS", "604800")),

        SMTP_HOST=os.getenv("SMTP_HOST"),
        SMTP_PORT=int(os.getenv("SMTP_PORT", "587")),
        SMTP_USE_TLS=str(os.getenv("SMTP_USE_TLS", "1")).lower() in {"1", "true", "yes"},
        SMTP_USER=os.getenv("SMTP_USER"),
        SMTP_PASS=os.getenv("SMTP_PASS"),
        SMTP_FROM=os.getenv("SMTP_FROM"),
        SMTP_FROM_NAME=os.getenv("SMTP_FROM_NAME"),
    )

_init_config_from_env(app)
# 환경변수 FLASK_SECRET_KEY 없으면 개발용 기본값 사용
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-me")
app.permanent_session_lifetime = timedelta(days=7)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True

print("TEMPLATE SEARCH PATH =", app.jinja_loader.searchpath)

def _load_docs():
    if not DOCS_DB.exists():
        return []
    try:
        return json.loads(DOCS_DB.read_text("utf-8"))
    except Exception:
        return []

def _save_docs(items):
    DOCS_DB.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")

def _allowed_file(filename: str):
    return "." in filename

def _ascii_safe_filename(original: str) -> str:
    name, ext = os.path.splitext(original)
    norm = unicodedata.normalize("NFKD", name)
    ascii_name = "".join(ch for ch in norm if ord(ch) < 128).strip()
    ascii_name = secure_filename(ascii_name) or uuid.uuid4().hex[:8]
    return ascii_name + ext  # 확장자는 유지

def _comma_split(s):
    # "a@a.com, b@b.com" -> ["a@a.com","b@b.com"]
    if not s:
        return []
    return [x.strip() for x in s.split(",") if x.strip()]

import io, zipfile, mimetypes, uuid, os, re
from datetime import datetime
from html import escape
from email.message import EmailMessage
import smtplib

MAX_PER_FILE = 20 * 1024 * 1024    # 20MB
MAX_TOTAL    = 24 * 1024 * 1024    # 메일 첨부 총합 안전 한도(25MB보다 살짝 낮게)

def _send_email_with_attachments(to_emails, subject, text_body, html_body=None, attachments=None):
    """
    attachments: list of tuples -> (filename, content_type, bytes)
    SMTP 설정은 기존에 쓰던 값 사용: SMTP_HOST, SMTP_PORT, SMTP_STARTTLS, SMTP_USERNAME, SMTP_PASSWORD, MAIL_FROM
    """
    msg = EmailMessage()
    msg["From"] = current_app.config.get("MAIL_FROM", "no-reply@jangbion.com")
    msg["To"]   = ", ".join(to_emails)
    msg["Subject"] = subject
    msg.set_content(text_body or "")
    if html_body:
        msg.add_alternative(html_body, subtype="html")

    for fname, ctype, data in attachments or []:
        maintype, _, subtype = (ctype or "application/octet-stream").partition("/")
        msg.add_attachment(
            data,
            maintype=maintype or "application",
            subtype=subtype or "octet-stream",
            filename=fname
        )

    host = current_app.config["SMTP_HOST"]
    port = int(current_app.config.get("SMTP_PORT", 587))
    with smtplib.SMTP(host, port) as s:
        if current_app.config.get("SMTP_STARTTLS", True):
            s.starttls()
        if current_app.config.get("SMTP_USERNAME"):
            s.login(current_app.config["SMTP_USERNAME"], current_app.config["SMTP_PASSWORD"])
        s.send_message(msg)

# === 메일 전송(첨부 지원) ===
def send_email_with_attachments(to_emails, subject, body_text, cc_emails=None, bcc_emails=None, attachments=None):
    """
    attachments: [(filename, bytes, mime)] 형태 리스트
    """
    to_emails = to_emails or []
    cc_emails = cc_emails or []
    bcc_emails = bcc_emails or []
    attachments = attachments or []

    # SMTP_* 는 이미 .env에서 읽어오는 걸 사용 중이라고 가정(이전 작업)
    SMTP_HOST = os.environ.get("SMTP_HOST")
    SMTP_PORT = int(os.environ.get("SMTP_PORT","587"))
    def _truthy(v): return str(v).lower() in ("1","true","yes","on","y")
    SMTP_USE_TLS = _truthy(os.environ.get("SMTP_USE_TLS", "1"))
    SMTP_USER = os.environ.get("SMTP_USER")
    SMTP_PASS = os.environ.get("SMTP_PASS")
    SMTP_FROM = os.environ.get("SMTP_FROM") or SMTP_USER
    SMTP_FROM_NAME = os.environ.get("SMTP_FROM_NAME") or "Our App"

    if not SMTP_HOST:
        raise RuntimeError("SMTP_HOST가 설정되지 않았습니다.")

    msg = EmailMessage()
    # 안전하게 UTF-8 헤더
    msg["Subject"] = str(Header(subject or "", "utf-8"))
    msg["From"] = formataddr((str(Header(SMTP_FROM_NAME, "utf-8")), SMTP_FROM))
    msg["To"] = ", ".join(to_emails)
    if cc_emails:
        msg["Cc"] = ", ".join(cc_emails)
    if bcc_emails:
        # BCC는 헤더에 표시 안하는 게 일반적이지만 EmailMessage는 Bcc 헤더를 보내지 않음.
        pass

    # 본문(텍스트)
    msg.set_content(body_text or "", subtype="plain", charset="utf-8")

    # 첨부
    for (fn, data, mime) in attachments:
        main_type, sub_type = (mime.split("/", 1) if (mime and "/" in mime) else ("application", "octet-stream"))
        msg.add_attachment(
            data,
            maintype=main_type,
            subtype=sub_type,
            filename=str(Header(fn, "utf-8"))
        )

    # 실제 발송
    import smtplib
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
        s.ehlo()
        if SMTP_USE_TLS:
            try:
                s.starttls(); s.ehlo()
            except Exception:
                pass
        if SMTP_USER and SMTP_PASS:
            s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg, from_addr=SMTP_FROM, to_addrs=to_emails + cc_emails + bcc_emails)
    return True

# SMTP(메일 본문에 '링크'만 전송)
SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_USE_TLS = str(os.environ.get("SMTP_USE_TLS", "1")).lower() not in ("0","false","no")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USER or "no-reply@example.com")
SMTP_FROM_NAME = os.environ.get("SMTP_FROM_NAME", "장비ON")

# 공유 링크 만료 기본일(일수)
DEFAULT_SHARE_EXPIRE_DAYS = int(os.environ.get("DEFAULT_SHARE_EXPIRE_DAYS", "7"))

def _docs_read(company: str) -> list:
    return load_json("documents.json", {}).get(company, [])

def _slug_for_s3(text: str, fallback: str = "company") -> str:
    """S3 키용: ASCII만 남기고 나머지는 - 로 치환"""
    if not text:
        return fallback
    s = unicodedata.normalize("NFKD", text)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9._-]+", "-", s).strip("-")
    return s or fallback

def _docs_write(company: str, rows: list):
    db = load_json("documents.json", {})
    db[company] = rows
    save_json("documents.json", db)

def _ensure_company_docs_dir(company: str) -> Path:
    p = DOCS_DIR / company
    p.mkdir(parents=True, exist_ok=True)
    return p

def _docs_db_read() -> dict:
    return load_json("documents.json", {})

def _docs_db_write(db: dict):
    save_json("documents.json", db)

def _docs_list(company: str) -> list[dict]:
    return list(_docs_db_read().get(company, []))

def _cloud_backend():
    return (os.environ.get("CLOUD_BACKEND") or "").strip().lower()

def _cloud_expire_seconds():
    v = os.environ.get("S3_EXPIRE_SECONDS") or os.environ.get("GCS_EXPIRE_SECONDS") or "604800"
    try: return max(60, int(v))
    except: return 604800

def _s3_client():
    import boto3
    region = os.environ.get("S3_REGION") or None
    return boto3.client("s3", region_name=region)

def _gcs_client():
    from google.cloud import storage
    return storage.Client()

def _cloud_key(company: str, stored_name: str) -> str:
    # 회사별 디렉터리로 분리
    return f"{company.strip()}/{stored_name}"

def _cloud_put(company: str, stored_name: str, data: bytes, mime: str) -> dict:
    """
    클라우드(또는 로컬)에 업로드하고 접근 URL을 돌려준다.
    return: {"where": "s3|gcs|local", "key": "...", "url": "..."}  (url은 즉시 클릭 가능한 링크)
    """
    backend = _cloud_backend()
    company = company or "default"

    if backend == "s3" and os.environ.get("S3_BUCKET"):
        bucket = os.environ["S3_BUCKET"]
        key = _cloud_key(company, stored_name)
        s3 = _s3_client()
        s3.put_object(Bucket=bucket, Key=key, Body=data, ContentType=(mime or "application/octet-stream"))
        # presigned GET url
        url = s3.generate_presigned_url(
            ClientMethod="get_object",
            Params={"Bucket": bucket, "Key": key},
            ExpiresIn=_cloud_expire_seconds(),
        )
        return {"where":"s3", "key": key, "url": url}

    if backend == "gcs" and os.environ.get("GCS_BUCKET"):
        bucket_name = os.environ["GCS_BUCKET"]
        client = _gcs_client()
        bucket = client.bucket(bucket_name)
        key = _cloud_key(company, stored_name)
        blob = bucket.blob(key)
        blob.upload_from_string(data, content_type=(mime or "application/octet-stream"))
        url = blob.generate_signed_url(expiration=_cloud_expire_seconds(), method="GET")
        return {"where":"gcs", "key": key, "url": url}

    # fallback: 로컬 저장(기존 방식)
    dest_dir = _ensure_company_docs_dir(company)
    path = dest_dir / stored_name
    path.write_bytes(data)
    # 로컬 파일은 /docs/file/<fname>로 내려주던 기존 라우트를 사용
    return {"where":"local", "key": stored_name, "url": url_for("docs_file", fname=stored_name, _external=True)}

def _docs_add(company: str, filename: str, stored: str, size: int, mime: str) -> dict:
    rec = {
        "id": uuid.uuid4().hex,
        "filename": filename,
        "stored": stored,
        "size": int(size or 0),
        "mime": mime or "application/octet-stream",
        "uploaded_at": _now_str(),
    }
    db = _docs_db_read()
    db.setdefault(company, []).append(rec)
    _docs_db_write(db)
    return rec

def _docs_find(company: str, key: str) -> dict | None:
    """id 우선, 없으면 stored/filename 순으로 검색"""
    for r in _docs_list(company):
        if r.get("id") == key or r.get("stored") == key or r.get("filename") == key:
            return r
    return None

def _docs_delete(company: str, key: str) -> bool:
    db = _docs_db_read()
    arr = list(db.get(company, []))
    tgt = None
    for r in arr:
        if r.get("id") == key or r.get("stored") == key or r.get("filename") == key:
            tgt = r; break
    if not tgt:
        return False
    # 파일 삭제
    p = _ensure_company_docs_dir(company) / tgt["stored"]
    try:
        if p.exists():
            p.unlink()
    except Exception:
        pass
    # 메타에서 제거
    arr = [r for r in arr if r.get("id") != tgt["id"]]
    db[company] = arr
    _docs_db_write(db)
    return True


def _parse_emails(s: str):
    if not s:
        return []
    parts = [p.strip() for p in s.replace(";", ",").split(",")]
    return [p for p in parts if "@" in p]

def _sendgrid_send(to_emails: list, subject: str, html_body: str, text_body: str = None):
    """
    SendGrid REST API(v3)로 메일을 전송한다. 환경변수:
      - SENDGRID_API_KEY
      - EMAIL_FROM
    """
    api_key = (os.environ.get("SENDGRID_API_KEY") or "").strip()
    from_email = (os.environ.get("EMAIL_FROM") or "").strip()

    if not api_key:
        raise RuntimeError("SENDGRID_API_KEY가 설정되지 않았습니다.")
    if not from_email:
        raise RuntimeError("EMAIL_FROM이 설정되지 않았습니다.")
    if not to_emails:
        raise RuntimeError("수신자(to)가 비어 있습니다.")

    # SendGrid payload
    payload = {
        "personalizations": [{
            "to": [{"email": x} for x in to_emails]
        }],
        "from": {"email": from_email, "name": "장비ON"},
        "subject": subject,
        "content": []
    }
    if text_body:
        payload["content"].append({"type": "text/plain", "value": text_body})
    payload["content"].append({"type": "text/html", "value": html_body})

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        },
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            body = r.read().decode("utf-8", "ignore")
            return r.getcode(), body
    except urllib.error.HTTPError as e:
        # SendGrid는 성공 시 202를 돌려준다.
        return e.code, e.read().decode("utf-8", "ignore")
    except Exception as e:
        return 0, str(e)

def _doc_find_by_id(company: str, doc_id: str):
    for d in _docs_read(company):
        if str(d.get("id")) == str(doc_id):
            return d
    return None

def _human_size(n: int) -> str:
    try:
        n = int(n)
    except Exception:
        return str(n)
    for unit in ("B","KB","MB","GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.0f}{unit}"
        n /= 1024.0

# ===== 공유 링크(패키지) 저장소 =====
def _shares_read_all() -> dict:
    return load_json("shares.json", {})

def _shares_write_all(data: dict):
    save_json("shares.json", data)

def _shares_read(company: str) -> list:
    return _shares_read_all().get(company, [])

def _shares_write(company: str, rows: list):
    db = _shares_read_all()
    db[company] = rows
    _shares_write_all(db)

def _find_share_by_token(token: str):
    db = _shares_read_all()
    for company, arr in db.items():
        for s in arr:
            if s.get("token") == token:
                return company, s
    return None, None

def _make_token(nbytes: int = 18) -> str:
    # URL-safe, 짧고 강력
    return secrets.token_urlsafe(nbytes)

def _hash_password(pw: str) -> str:
    # 간단 SHA-256 + salt
    salt = secrets.token_hex(8)
    h = hashlib.sha256((salt + pw).encode("utf-8")).hexdigest()
    return f"{salt}:{h}"

def _to_list(v):
    if not v:
        return []
    if isinstance(v, (list, tuple, set)):
        return [str(x).strip() for x in v if str(x).strip()]
    # "a@b.com, c@d.com ; e@f.com" 같은 문자열도 허용
    return [s.strip() for s in re.split(r"[;,]", str(v)) if s.strip()]

def _check_password(pw: str, hashed: str) -> bool:
    try:
        salt, h = hashed.split(":", 1)
        return hashlib.sha256((salt + pw).encode("utf-8")).hexdigest() == h
    except Exception:
        return False

def _send_email_link(
    to_emails,
    subject: str,
    body_text: str,
    cc_emails=None,
    bcc_emails=None,
    body_html: str | None = None,
):
    to_emails = to_emails or []
    cc_emails = cc_emails or []
    bcc_emails = bcc_emails or []

    msg = EmailMessage()
    msg["Subject"] = str(Header(subject or "", "utf-8"))
    msg["From"] = formataddr((str(Header(SMTP_FROM_NAME, "utf-8")), SMTP_FROM))
    msg["To"] = ", ".join(to_emails)
    if cc_emails:
        msg["Cc"] = ", ".join(cc_emails)
    # Bcc는 헤더에 넣지 않음

    # 본문
    msg.set_content(body_text or "", subtype="plain", charset="utf-8")
    if body_html:
        msg.add_alternative(body_html, subtype="html")  # utf-8 기본 적용

    all_rcpts = to_emails + cc_emails + bcc_emails

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
        s.ehlo()
        if SMTP_USE_TLS:
            try:
                s.starttls(); s.ehlo()
            except Exception:
                pass
        if SMTP_USER and SMTP_PASS:
            s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg, from_addr=SMTP_FROM, to_addrs=all_rcpts)


FINANCE_FILTER_KEYS = [
    "tab","start","end","page","page_size","half",
    "by_worker","by_machine","by_client",
    "status","pay","unpaid_only",
    "worker","worker_input","plate","owner","tenant",
    "inc_cat","inc_desc","exp_cat","exp_desc",
]
def _carry_params_from_form():
    return {k: v for k, v in (request.form or {}).items()
            if k in FINANCE_FILTER_KEYS and str(v).strip() != ""}

def _ensure_company_bucket(db_any, company: str):
    # 파일이 리스트 형태로 저장돼 있던 레거시도 안전 처리
    if isinstance(db_any, list):
        db = {company: list(db_any)}
    elif isinstance(db_any, dict):
        db = dict(db_any)
    else:
        db = {}
    if company not in db or not isinstance(db.get(company), list):
        db[company] = []
    return db

# ──[ jobs → 자동 수입/지출 동기화용 안정 해시키 ]────────────────────
import hashlib, json as _json

def _stable_job_key(company: str, j: dict) -> str:
    """jobs.json의 한 건을 안정적으로 식별하는 해시 키(중복생성 방지)"""
    basis = {
        "company": company,
        "date": (j.get("date") or "").strip(),
        "worker": (j.get("worker") or j.get("driver") or "").strip(),
        "machine_number": (j.get("machine_number") or j.get("plate") or "").strip(),
        "owner": (j.get("client_primary") or j.get("client") or "").strip(),
        "tenant": (j.get("client_tenant") or "").strip(),
        "outsource_type": (j.get("outsource_type") or "none").strip(),
        "amount_man": int(str(j.get("amount_man") or 0).replace(",", "") or 0),
    }
    raw = _json.dumps(basis, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]

@app.before_request
def _dev__always_reload_templates():
    app.jinja_env.cache.clear()

# =========================================================
# 공통 유틸(단일 정의, 중복 제거)
# =========================================================
KEY_ORDER = (
    "amount","price","total","sum","revenue","income","fee","payment","paid","receivable",
    "금액","가격","총액","합계","청구","요금","수입","매출","대금"
)


def _company_name():
    return (current_app.config.get("COMPANY_NAME")
            or os.getenv("COMPANY_NAME")
            or "company")

def _out_key_for_job(j: dict, kind: str) -> str:
    """_sync_outsourcing_entries()에서 만드는 auto_key와 동일 규칙으로 키 생성"""
    base = "|".join([
        str(j.get("date","")),
        str(j.get("worker","")),
        str(j.get("machine_number","")),
        str(j.get("outsource_partner","")),
        kind.lower().strip()
    ])
    # hashlib는 이미 상단에 import 되어 있음
    return "auto-" + hashlib.md5(base.encode("utf-8")).hexdigest()[:16]

def load_json(filename, default):
    p = DATA_DIR / filename
    try:
        with p.open("r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception:
        return default

def save_json(filename, data):
    p = DATA_DIR / filename
    tmp = p.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, p)

def _to_number(x: object) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        if not s: return 0.0
        m = re.search(r"-?\d+(?:[,\s]?\d{3})*(?:\.\d+)?", s)
        if m:
            try:
                return float(m.group(0).replace(",", "").replace(" ", ""))
            except ValueError:
                return 0.0
    return 0.0

def _now_str() -> str:
    try:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

def _parse_date_safe(s):
    if isinstance(s, date):
        return s
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    if 'T' in s:
        s = s.split('T', 1)[0]
    if ' ' in s:
        s = s.split(' ', 1)[0]
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return date.fromisoformat(s)
    except Exception:
        return None

def _in_range(d: date, start: date, end: date):
    return d is not None and (not start or d >= start) and (not end or d <= end)

def _amount_won(job: dict) -> int:
    """jobs.json의 금액(만원)을 원으로"""
    if not isinstance(job, dict):
        return 0
    for k in ("amount","total","sum","price","금액","총액","합계","청구금액","합계금액","계약금액"):
        if k in job:
            v = int(_to_number(job.get(k)))
            if v: return v
    for k in ("amount_man","total_man","paid_amount_man_total"):
        if k in job:
            v = int(_to_number(job.get(k)))
            if v: return v * 10000
    for a, b in [("unit_price","qty"), ("rate","hours"), ("단가","수량"), ("단가","시간"), ("일당","일수")]:
        if a in job and b in job:
            na, nb = _to_number(job[a]), _to_number(job[b])
            if na and nb: return int(na * nb)
    return 0

def _paid_won(job: dict) -> int:
    for k in ("paid","paid_amount","received","수납","입금","총지급","실지급"):
        if k in job:
            v = int(_to_number(job.get(k)))
            if v: return v
    for k in ("paid_amount_man","paid_man","paid_total_man"):
        if k in job:
            v = int(_to_number(job.get(k)))
            if v: return v * 10000
    return 0

def _color_by_payment(amount_won: int, paid_won: int, status: str) -> str:
    s = (status or "").strip()
    if any(x in s for x in ("완납", "완불")) or (amount_won and paid_won >= amount_won):
        return "paid"
    if "부분" in s:
        return "partial"
    if "미완" in s:
        return "unpaid"
    if amount_won > 0:
        if paid_won == 0:
            return "unpaid"
        if 0 < paid_won < amount_won:
            return "partial"
    return "paid"

# 표시용: 작업 금액 계산 (템플릿에서 사용)
def job_amount(j: dict) -> float:
    if not isinstance(j, dict): return 0.0
    for k in KEY_ORDER:
        if k in j:
            n = _to_number(j.get(k))
            if n != 0:
                if k.endswith("_man"):  # 만원단위 키
                    return n * 10000.0
                return n
    for a, b in [("unit_price","qty"), ("rate","hours"), ("단가","수량"), ("단가","시간"), ("일당","일수")]:
        if a in j and b in j:
            na, nb = _to_number(j[a]), _to_number(j[b])
            if na and nb:
                return na * nb
    for v in j.values():
        if isinstance(v, dict):
            n = job_amount(v)
            if n: return n
    return 0.0

app.add_template_global(job_amount, name="job_amount")

# =========================================================
# 수입/지출 IO
# =========================================================
def incomes_read(company: str) -> list:
    return load_json("incomes.json", {}).get(company, [])

def incomes_write(company: str, rows: list):
    db = load_json("incomes.json", {})
    db[company] = rows
    save_json("incomes.json", db)

def expenses_read(company: str) -> list:
    return load_json("expenses_db.json", {}).get(company, [])

def expenses_write(company: str, rows: list):
    db = load_json("expenses_db.json", {})
    db[company] = rows
    save_json("expenses_db.json", db)

def _check_s3_config_and_flash():
    cfg = current_app.config
    if (cfg.get("CLOUD_BACKEND") or "").lower() != "s3":
        return
    missing = []
    for k in ("S3_BUCKET","S3_REGION"):
        if not (cfg.get(k) or os.getenv(k)):
            missing.append(k)
    if not (os.getenv("AWS_ACCESS_KEY_ID") and os.getenv("AWS_SECRET_ACCESS_KEY")):
        for k in ("AWS_ACCESS_KEY_ID","AWS_SECRET_ACCESS_KEY"):
            if not os.getenv(k):
                missing.append(k)
    if missing:
        flash(f"S3 설정이 누락되었습니다: {', '.join(sorted(set(missing)))} (.env 확인)", "error")

# =========================================================
# 외주 동기화(중복 방지·정리) — 결과 요약 반환으로 개선
# =========================================================
def _sync_outsourcing_entries(company, jobs):
    incomes  = load_json("incomes.json", {})
    expenses = load_json("expenses_db.json", {})
    incomes.setdefault(company, [])
    expenses.setdefault(company, [])

    want_inc, want_exp = {}, {}

    for j in (jobs or []):
        otype   = str(j.get("outsource_type","")).lower()  # 'received' | 'given' | 'none'
        partner = (j.get("outsource_partner") or "").strip()
        if not partner or otype in ("", "none"):
            continue

        amount = job_out_amount_won(j)  # 만원 → 원 환산 포함
        key = outsrc_auto_key(j, otype)

        item = {
            "id": key,
            "auto_key": key,
            "source": "auto_outsrc_received" if otype == "received" else "auto_outsrc_given",
            "date": j.get("date",""),
            "category": "외주받음" if otype == "received" else "외주줬음",
            "desc": partner,
            "amount": amount,
        }

        if otype == "received":
            want_inc[key] = item
        else:
            want_exp[key] = item

    def sync_list(cur_list, want_map):
        # 기존 자동항목 인덱싱
        idx = { (x.get("auto_key") or x.get("id")): i for i, x in enumerate(cur_list) if (x.get("auto_key") or x.get("id")) }
        added = removed = 0
        keep = []
        for x in cur_list:
            k = x.get("auto_key") or x.get("id")
            if str(x.get("source","")).startswith("auto_outsrc_") and (k not in want_map):
                removed += 1
                continue
            keep.append(x)
        for k, v in want_map.items():
            if k not in idx:
                keep.append(v); added += 1
        return keep, added, removed

    def sync_list(cur_list, want_map):
        idx = {}
        for i, x in enumerate(cur_list):
            k = x.get("auto_key") or x.get("id")
            if k: idx[k] = i

        added = removed = 0
        keep = []
        for x in cur_list:
            k = x.get("auto_key") or x.get("id")
            if x.get("source","").startswith("auto_outsrc_") and (k not in want_map):
                removed += 1
                continue
            keep.append(x)

        for k, v in want_map.items():
            if k not in idx:
                keep.append(v); added += 1
        return keep, added, removed

    incomes[company],  inc_added, inc_removed = sync_list(incomes[company],  want_inc)
    expenses[company], exp_added, exp_removed = sync_list(expenses[company], want_exp)

    save_json("incomes.json", incomes)
    save_json("expenses_db.json", expenses)
    return {
        "inc_added": inc_added, "inc_removed": inc_removed,
        "exp_added": exp_added, "exp_removed": exp_removed
    }

# =========================================================
# 템플릿 필터/유틸
# =========================================================
@app.template_filter("won")
def won(v):
    try:
        return f"{int(round(float(v))):,}원"
    except Exception:
        return v

def get_current_user():
    users = load_json('users.json', {})
    return users.get(session.get('username'), {}) or {}

def is_admin(user=None):
    if user is None:
        user = get_current_user()
    return (user.get('role') in ('boss', 'manager'))

def has_perm(perm: str) -> bool:
    role = session.get('role')
    PERMISSIONS = {
        'boss': {'view_dashboard','manage_clients','manage_jobs','manage_workers','manage_machines','approve_workers','manage_roles','manage_payments'},
        'manager': {'view_dashboard','manage_clients','manage_jobs','manage_workers','manage_machines','manage_payments'},
        'worker': {'view_dashboard_worker','view_my_jobs','update_job_status'}
    }
    return perm in PERMISSIONS.get(role, set())

def perm_required(*perms):
    def deco(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not session.get('username'):
                return redirect(url_for('login'))
            if not any(has_perm(p) for p in perms):
                return redirect(url_for('dashboard_worker'))
            return f(*args, **kwargs)
        return wrapped
    return deco

def redirect_with_from(endpoint, **kwargs):
    src = request.args.get('from') or request.form.get('from')
    if src:
        kwargs['from'] = src
    return redirect(url_for(endpoint, **kwargs))

def back_with_error(*args, **kwargs):
    if len(args) == 1 and isinstance(args[0], str):
        msg = args[0]
        return f"<script>alert({json.dumps(msg, ensure_ascii=False)});history.back();</script>"
    if len(args) >= 2:
        endpoint, message = args[0], args[1]
        return redirect_with_from(endpoint, error=message, **kwargs)
    return redirect(url_for('home'))

def jinja_split(value, sep=':', maxsplit=-1):
    try:
        return (value or '').split(sep, maxsplit)
    except Exception:
        return []
app.jinja_env.filters['split'] = jinja_split
app.jinja_env.globals['str'] = str
app.jinja_env.filters['string'] = lambda x: '' if x is None else str(x)

# 파트너 IO
def load_partners(company: str):
    data = load_json('partners.json', {})
    entry = data.get(company)
    if not entry:
        legacy = load_json('clients.json', {}).get(company, [])
        entry = {"owners": legacy[:], "tenants": []}
        data[company] = entry
        save_json('partners.json', data)
    return entry

def save_partners(company: str, owners: list, tenants: list):
    data = load_json('partners.json', {})
    data[company] = {"owners": owners, "tenants": tenants}
    save_json('partners.json', data)

def outsrc_auto_key(job: dict, kind: str) -> str:
    """외주(받음/줬음) 자동 수입/지출 항목의 안정 키"""
    base = "|".join([
        str(job.get("date", "")),
        str(job.get("worker", "")),
        str(job.get("machine_number", "")),
        str(job.get("outsource_partner", "")),
        kind.lower().strip()
    ])
    import hashlib as _h
    return "auto-" + _h.md5(base.encode("utf-8")).hexdigest()[:16]

def _delete_outsourcing_job_by_auto_item(company: str, auto_item: dict) -> bool:
    """
    자동 외주 수입/지출(auto_outsrc_*) 항목을 받아서,
    대응하는 jobs.json의 원본 작업을 찾아 삭제한다.
    찾으면 True, 없으면 False.
    """
    if not auto_item or not str(auto_item.get("source","")).startswith("auto_outsrc_"):
        return False

    kind = "received" if auto_item["source"] == "auto_outsrc_received" else "given"
    auto_key = auto_item.get("auto_key") or auto_item.get("id") or ""

    jobs_db = load_json('jobs.json', {})
    job_list = list(jobs_db.get(company, []))

    # 1) auto_key로 정밀 매칭
    for i, j in enumerate(job_list):
        try:
            if (j.get("outsource_type","").lower() == kind
                and outsrc_auto_key(j, kind) == auto_key):
                job_list.pop(i)
                jobs_db[company] = job_list
                save_json('jobs.json', jobs_db)
                return True
        except Exception:
            pass

    # 2) 필드 기반 보강 매칭 (날짜+파트너+종류+금액)
    tgt_date   = (auto_item.get("date") or "").strip()
    tgt_partner= (auto_item.get("desc") or "").strip()
    tgt_amount = int(_to_number(auto_item.get("amount") or 0))

    for i, j in enumerate(job_list):
        if j.get("outsource_type","").lower() != kind:
            continue
        if (j.get("date") or "").strip() != tgt_date:
            continue
        if (j.get("outsource_partner") or "").strip() != tgt_partner:
            continue
        if job_out_amount_won(j) != tgt_amount:
            continue
        # 충분히 동일하다고 판단
        job_list.pop(i)
        jobs_db[company] = job_list
        save_json('jobs.json', jobs_db)
        return True

    return False

def job_out_amount_won(job: dict) -> int:
    """외주 금액(원): out_amount > out_amount_man*1만 > amount_man*1만"""
    if job.get("out_amount") not in (None, ""):
        return int(_to_number(job.get("out_amount")))
    if job.get("out_amount_man") not in (None, ""):
        return int(_to_number(job.get("out_amount_man"))) * 10000
    return int(_to_number(job.get("amount_man") or 0)) * 10000

def _seed_company_containers(company):
    for fname, seed in [
        ("workers.json", []),
        ("machines.json", []),
        ("clients.json", []),
        ("partners.json", {"owners": [], "tenants": []}),
        ("jobs.json", []),
    ]:
        db = load_json(fname, {})
        if company not in db:
            db[company] = seed
            save_json(fname, db)

# =========================================================
# 인증/대시보드
# =========================================================
@app.route('/')
def home():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        phone = request.form.get('phone')
        password = request.form.get('password')

        users_db = load_json('users.json', {})
        user = None
        username = None
        for u_name, u_data in users_db.items():
            if u_data.get('phone') == phone:
                user = u_data
                username = u_name
                break

        if user and user.get('password') == password:
            if user.get('role') == 'worker' and user.get('status', 'active') == 'pending':
                error = '승인 대기중입니다. 사장님의 승인을 기다려주세요.'
                return render_template('login.html', error=error)

            session.permanent = True
            session['username'] = username
            session['role'] = user.get('role')
            session['company'] = user.get('company', '')

            if user['role'] in ('boss', 'manager'):
                return redirect(url_for('dashboard'))
            return redirect(url_for('dashboard_worker'))
        else:
            error = '휴대폰번호 또는 비밀번호가 올바르지 않습니다.'
    return render_template('login.html', error=error)

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    if not is_admin():
        return redirect(url_for('dashboard_worker'))
    return render_template('dashboard.html')

@app.route('/dashboard_worker')
def dashboard_worker():
    if 'username' not in session:
        return redirect(url_for('login'))
    if is_admin():
        return redirect(url_for('dashboard'))
    return render_template('dashboard_worker.html')

# =========================================================
# 작업 목록/등록/수정/삭제/엑셀/결제/캘린더 (원기능 유지)
# =========================================================
@app.route('/jobs', endpoint='view_jobs')
@app.route('/view_jobs', endpoint='view_jobs_legacy')
def view_jobs():
    import math
    if 'username' not in session:
        return redirect(url_for('login'))

    user = get_current_user()
    if not is_admin(user):
        return redirect(url_for('dashboard_worker'))

    company = user.get('company', '')
    jobs_all = load_json('jobs.json', {}).get(company, [])
    for i, j in enumerate(jobs_all):
        j['_idx'] = i

    partners = load_partners(company)
    owners_list  = partners.get('owners', [])
    tenants_list = partners.get('tenants', [])

    # ---------- 쿼리 ----------
    q_worker = (request.args.get('worker') or '').strip()
    q_owner  = (request.args.get('owner') or request.args.get('client_primary') or request.args.get('client') or '').strip()
    q_tenant = (request.args.get('tenant') or request.args.get('client_tenant') or '').strip()
    q_date   = (request.args.get('date') or '').strip()
    q_from   = (request.args.get('date_from') or '').strip()
    q_to     = (request.args.get('date_to') or '').strip()

    raw_status = (request.args.get('status') or '').strip()
    def _norm_status(s: str) -> str:
        s = s.strip()
        if s in ('pending','진행중','todo','ing'): return 'pending'
        if s in ('done','완료','complete'):        return 'done'
        return ''
    status_filter = _norm_status(raw_status)

    pay_filter = (request.args.get('pay') or '').strip()
    spare  = (request.args.get('spare')  or '').strip()
    outsrc = (request.args.get('outsrc') or '').strip()

    overdue = (request.args.get('overdue') or '').strip()
    dues    = (request.args.get('dues')    or '').strip()
    if overdue == '1' and not status_filter:
        status_filter = 'pending'
    if dues == '1' and not pay_filter:
        pay_filter = 'unpaid'

    def _contains(hay, needle):
        return (needle == '' or (hay or '').lower().find(needle.lower()) != -1)

    jobs = jobs_all
    if q_worker:
        jobs = [j for j in jobs if (j.get('worker') or '') == q_worker]
    if q_owner:
        jobs = [j for j in jobs if _contains(j.get('client_primary') or j.get('client'), q_owner)]
    if q_tenant:
        jobs = [j for j in jobs if _contains(j.get('client_tenant'), q_tenant)]
    if q_from or q_to:
        def in_range(d):
            if not d: return False
            return (not q_from or d >= q_from) and (not q_to or d <= q_to)
        jobs = [j for j in jobs if in_range((j.get('date') or '').strip())]
    elif q_date:
        jobs = [j for j in jobs if q_date == (j.get('date') or '')]

    if status_filter == 'pending':
        jobs = [j for j in jobs if (j.get('status') or '진행중').strip() != '완료']
    elif status_filter == 'done':
        jobs = [j for j in jobs if (j.get('status') or '').strip() == '완료']

    def _to_int(x):
        try: return int(x)
        except: return 0
    if pay_filter == 'unpaid':
        jobs = [j for j in jobs if _to_int(j.get('amount_man')) > 0 and _to_int(j.get('paid_amount_man')) < _to_int(j.get('amount_man'))]
    elif pay_filter == 'paid':
        jobs = [j for j in jobs if _to_int(j.get('amount_man')) > 0 and _to_int(j.get('paid_amount_man')) >= _to_int(j.get('amount_man'))]

    if spare == '1':
        jobs = [j for j in jobs if bool(j.get('is_spare'))]
    if outsrc == '1':
        jobs = [j for j in jobs if (j.get('outsource_type') or 'none') != 'none']

    def _dt_key(j):
        d = (j.get('date') or '').strip()
        t = (j.get('time') or '').strip() or "00:00"
        try:
            return _dt.datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
        except:
            try:
                return _dt.datetime.strptime(d, "%Y-%m-%d")
            except:
                return _dt.datetime.min
    def _sort_key(j):
        return (_dt_key(j), j.get('_idx', -1))
    jobs = sorted(jobs, key=_sort_key, reverse=True)

    per_page = int(request.args.get('per_page', 20) or 20)
    page = int(request.args.get('page', 1) or 1)
    total_count = len(jobs)
    pages = max(1, math.ceil(total_count / per_page)) if per_page > 0 else 1
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end   = start + per_page
    page_jobs = jobs[start:end]
    all_ids = [j['_idx'] for j in jobs]

    args_dict = request.args.to_dict(flat=True)
    def _build_cycle(param, next_val):
        q = args_dict.copy()
        q.pop('overdue', None); q.pop('dues', None)
        q.pop('page', None)
        if next_val: q[param] = next_val
        else: q.pop(param, None)
        return url_for('view_jobs', **q)

    if (status_filter or '') == '':
        status_label, status_next = '작업상태 필터', 'pending'
    elif status_filter == 'pending':
        status_label, status_next = '진행중인 작업만', 'done'
    else:
        status_label, status_next = '완료된 작업만', ''
    status_cycle_url = _build_cycle('status', status_next)

    if (pay_filter or '') == '':
        pay_label, pay_next = '지불상태 필터', 'unpaid'
    elif pay_filter == 'unpaid':
        pay_label, pay_next = '미납된 작업만', 'paid'
    else:
        pay_label, pay_next = '완납된 작업만', ''
    pay_cycle_url = _build_cycle('pay', pay_next)

    def _onoff(key):
        on  = args_dict.copy(); on[key] = '1'; on.pop('page', None)
        off = args_dict.copy(); off.pop(key, None); off.pop('page', None)
        on.pop('overdue', None); on.pop('dues', None)
        off.pop('overdue', None); off.pop('dues', None)
        return url_for('view_jobs', **on), url_for('view_jobs', **off)
    spare_on_url,  spare_off_url  = _onoff('spare')
    outsrc_on_url, outsrc_off_url = _onoff('outsrc')

    return render_template(
        'view_job.html',
        jobs=page_jobs,
        page=page, pages=pages, per_page=per_page, total_count=total_count,
        q_worker=q_worker, q_owner=q_owner, q_tenant=q_tenant,
        q_date=q_date, q_from=q_from, q_to=q_to,
        spare=spare, spare_on_url=spare_on_url, spare_off_url=spare_off_url,
        outsrc=outsrc, outsrc_on_url=outsrc_on_url, outsrc_off_url=outsrc_off_url,
        status_filter=status_filter, status_label=status_label, status_cycle_url=status_cycle_url,
        pay_filter=pay_filter,       pay_label=pay_label,       pay_cycle_url=pay_cycle_url,
        owners=owners_list, tenants=tenants_list,
        all_ids=all_ids
    )

@app.post('/export_selected_xlsx')
@app.post('/jobs/export/xlsx')
def export_selected_xlsx():
    from io import BytesIO
    import datetime as _dt
    from urllib.parse import quote
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl가 없습니다. venv에서 'pip install openpyxl' 실행 후 다시 시도하세요.", 500

    # 1) 선택 인덱스 수집
    idx_list = request.form.getlist("selected_jobs")
    indices, _seen = [], set()
    for x in idx_list:
        try:
            i = int(x)
        except (TypeError, ValueError):
            continue
        if i not in _seen:
            _seen.add(i)
            indices.append(i)
    if not indices:
        return "선택된 항목이 없습니다.", 400

    # 2) 컬럼 구성
    default_cols = [
        "date","owner","tenant","machine_name","machine_number","machine_alias",
        "worker","amount_full","payment_status","remaining_full",
        "status","location","note","outsource_recv","outsource_give","spare"
    ]
    cols_req = request.form.getlist("cols")
    cols = [c for c in cols_req if c in default_cols] or default_cols

    header_map = {
        "date":"날짜", "owner":"원수급자", "tenant":"임차인",
        "machine_name":"장비명", "machine_number":"차량번호", "machine_alias":"별칭",
        "worker":"기사",
        "amount_full":"금액(원)", "payment_status":"지불상태", "remaining_full":"잔액(원)",
        "status":"작업상태", "location":"위치", "note":"요청사항",
        "outsource_recv":"외주받음", "outsource_give":"외주줬음", "spare":"스페어"
    }

    # 3) 데이터 로드
    company = session.get("company")
    all_jobs = load_json('jobs.json', {}).get(company, [])
    pick = [all_jobs[i] for i in indices if 0 <= i < len(all_jobs)]

    # 기간(파일명용): 폼에서 start/end가 오면 사용, 없으면 선택 항목의 최소/최대 날짜 사용
    def _safe_date(s):
        return _parse_date_safe((s or "").strip())

    req_start = _safe_date(request.form.get("start"))
    req_end   = _safe_date(request.form.get("end"))

    dates = [ _safe_date(j.get("date")) for j in pick if _safe_date(j.get("date")) ]
    min_d = min(dates) if dates else req_start or _dt.date.today()
    max_d = max(dates) if dates else req_end   or _dt.date.today()

    start_code = (req_start or min_d).strftime('%Y%m%d')
    end_code   = (req_end   or max_d).strftime('%Y%m%d')

    # 4) 정렬
    def _dt_key(j):
        d = (j.get('date') or '').strip()
        t = (j.get('time') or '').strip() or "00:00"
        try:
            return _dt.datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
        except Exception:
            try:
                return _dt.datetime.strptime(d, "%Y-%m-%d")
            except Exception:
                return _dt.datetime.min
    pick.sort(key=_dt_key, reverse=True)

    # 5) 행 생성
    def row_from_job(j):
        amount_man = int(j.get("amount_man") or 0)
        paid_man   = int(j.get("paid_amount_man") or 0)
        amount_full    = amount_man * 10000
        paid_full      = paid_man   * 10000
        remaining_full = max(amount_full - paid_full, 0)

        pstatus = (j.get("payment_status") or "").strip()
        if not pstatus:
            if amount_man <= 0: pstatus = "미설정"
            elif paid_man >= amount_man: pstatus = "완납"
            elif paid_man <= 0: pstatus = "미납"
            else: pstatus = "부분"

        ot = (j.get("outsource_type") or "none").strip()
        partner = j.get("outsource_partner") or ""
        recv = partner if ot == "received" else ""
        give = partner if ot == "given"    else ""

        mapping = {
            "date": j.get("date") or "",
            "owner": (j.get("client_primary") or j.get("client") or ""),
            "tenant": j.get("client_tenant") or "",
            "machine_name": j.get("machine_name") or j.get("machine") or "",
            "machine_number": j.get("machine_number") or "",
            "machine_alias": j.get("machine_alias") or "",
            "worker": j.get("worker") or "",
            "amount_full": amount_full,
            "payment_status": pstatus,
            "remaining_full": remaining_full,
            "status": j.get("status") or "",
            "location": j.get("location") or "",
            "note": j.get("note") or "",
            "outsource_recv": recv,
            "outsource_give": give,
            "spare": "스페어" if j.get("is_spare") else "",
        }
        return [mapping[c] for c in cols]

    # 6) 엑셀 작성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "작업목록"

    # 머리글
    ws.append([header_map[c] for c in cols])
    header = ws[1]
    for c in header:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # 본문
    for j in pick:
        ws.append(row_from_job(j))

    # (요청사항) 화살표 제거: 자동필터 설정하지 않음 (아래 두 줄 모두 없음)
    # ws.auto_filter.ref = ws.dimensions
    # ws.auto_filter.add_filter_column(...)

    # 필요하면 헤더 고정만 유지
    ws.freeze_panes = "A2"

    # 숫자 열 서식/정렬
    head_to_col = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column+1)}
    right_cols = []
    for title in ("금액(원)", "잔액(원)"):
        col = head_to_col.get(title)
        if col:
            right_cols.append(col)
            for r in range(2, ws.max_row+1):
                cell = ws.cell(row=r, column=col)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # 나머지는 좌측 정렬
    for r in range(2, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            if c not in right_cols:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='left', vertical='center')

    # 테두리(지출엑셀과 동일한 얇은 테두리)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = thin

    # 자동 열 너비(내용 기준 + 최소너비 보정)
    def _autofit_worksheet(ws, min_widths=None, max_width=60):
        min_widths = min_widths or {}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                val = "" if cell.value is None else str(cell.value)
                # 줄바꿈 고려(가장 긴 줄 기준)
                max_len = max(max_len, max((len(x) for x in val.splitlines()), default=0))
            target = max(min_widths.get(col_idx, 10), min(max_width, max_len + 2))
            ws.column_dimensions[letter].width = target

    # 헤더명 -> 열번호 매핑으로 최소너비 지정
    minw = {}
    def set_min(header_text, w):
        col = head_to_col.get(header_text)
        if col: minw[col] = w

    set_min("날짜", 12)
    set_min("원수급자", 22)
    set_min("임차인", 22)
    set_min("장비명", 16)
    set_min("차량번호", 12)
    set_min("별칭", 10)
    set_min("기사", 12)
    set_min("금액(원)", 14)
    set_min("잔액(원)", 14)
    set_min("작업상태", 12)
    set_min("위치", 16)
    set_min("요청사항", 18)
    set_min("외주받음", 12)
    set_min("외주줬음", 12)
    set_min("스페어", 8)

    _autofit_worksheet(ws, min_widths=minw, max_width=70)

    # 7) 다운로드
    bio = BytesIO()
    wb.save(bio); bio.seek(0)

    filename = f"작업목록 {start_code}_{end_code}.xlsx"
    q = quote(filename)
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={q}; filename*=UTF-8''{q}"}
    )

@app.route('/add_job', methods=['GET', 'POST'])
@perm_required('manage_jobs')
def add_job():
    users_db = load_json('users.json', {})
    username = session.get('username')
    company = users_db.get(username, {}).get('company', '')

    workers   = load_json('workers.json',   {}).get(company, [])
    machines  = load_json('machines.json',  {}).get(company, [])
    locations = load_json('locations.json', {}).get(company, [])
    partners  = load_partners(company)
    owners    = partners.get('owners', [])
    tenants   = partners.get('tenants', [])

    if request.method == 'POST':
        time_str = (request.form.get('time') or '').strip()
        if not time_str:
            hour   = (request.form.get('hour')   or '').strip()
            minute = (request.form.get('minute') or '').strip()
            if hour and not minute: minute = '00'
            time_str = f"{hour}:{minute}" if hour and minute else ''
        date_str = (request.form.get('date') or '').strip()

        worker = (request.form.get('worker_input') or request.form.get('worker') or '').strip()

        machine_name   = (request.form.get('machine_name_input')   or request.form.get('machine_name')   or '').strip()
        machine_number = (request.form.get('machine_number_input') or request.form.get('machine_number') or '').strip()
        machine_alias  = (request.form.get('machine_alias_input')  or request.form.get('machine_alias')  or '').strip()

        client_primary = (request.form.get('client_primary') or request.form.get('client') or '').strip()
        client_tenant  = (request.form.get('client_tenant')  or '').strip()
        if request.form.get('same_as_owner'):
            client_tenant = client_primary

        if request.form.get('save_owner') == '1' and client_primary and client_primary not in owners:
            owners.append(client_primary)
        if request.form.get('save_tenant') == '1' and client_tenant and client_tenant not in tenants:
            tenants.append(client_tenant)
        save_partners(company, owners, tenants)

        location = (request.form.get('location_input') or request.form.get('location') or '').strip()
        note     = (request.form.get('note') or '').strip()

        duration_type  = (request.form.get('duration_type')  or '하루').strip()
        duration_hours = (request.form.get('duration_hours') or '').strip()
        if duration_type != 'N시간':
            duration_hours = ''

        amount_raw = (request.form.get('amount_man') or '').strip()
        try: amount_man = int(amount_raw) if amount_raw != '' else 0
        except ValueError: amount_man = 0
        share_amount = bool(request.form.get('share_amount'))

        outsource_type = (request.form.get('outsource_type') or 'none').strip()
        if outsource_type not in ('none', 'received', 'given'):
            outsource_type = 'none'
        outsource_partner = (request.form.get('outsource_partner') or '').strip()
        if outsource_type == 'none':
            outsource_partner = ''

        missing = []
        if not worker:         missing.append('기사')
        if not machine_name:   missing.append('장비명')
        if not machine_number: missing.append('차량번호')
        if not client_primary: missing.append('원수급자')
        if not location:       missing.append('위치')
        if not date_str:       missing.append('날짜')
        if not time_str:       missing.append('시간')
        if missing:
            return render_template(
                'add_job.html',
                workers=workers, machines=machines, locations=locations,
                owners=owners, tenants=tenants,
                error="다음 항목을 확인해 주세요: " + ", ".join(missing),
                prev=request.form,
                job_registered=False
            )

        payment_status  = '미설정' if amount_man <= 0 else '미납'
        paid_amount_man = 0

        new_job = {
            "date": date_str, "time": time_str,
            "worker": worker, "is_spare": bool(request.form.get('is_spare')),
            "machine_name": machine_name,
            "machine_number": machine_number,
            "machine_alias": machine_alias,
            "client_primary": client_primary,
            "client_tenant":  client_tenant,
            "client": client_primary,
            "location": location,
            "note": note,
            "status": "진행중",
            "duration_type": duration_type,
            "duration_hours": duration_hours,
            "amount_man": amount_man,
            "share_amount": share_amount,
            "outsource_type": outsource_type,
            "outsource_partner": outsource_partner,
            "payment_status":  payment_status,
            "paid_amount_man": paid_amount_man
        }

        try:
            locdb = load_json('locations.json', {})
            locs = locdb.get(company, [])
            if location:
                locs = [location] + [x for x in locs if x != location]
                locs = locs[:20]
                locdb[company] = locs
                save_json('locations.json', locdb)
        except Exception:
            pass

        jobs_db = load_json('jobs.json', {})
        jobs_db.setdefault(company, []).append(new_job)
        save_json('jobs.json', jobs_db)

        return render_template(
            'add_job.html',
            workers=workers, machines=machines, locations=locations,
            owners=owners, tenants=tenants,
            prev={}, job_registered=True,
            registered_filters={"date": new_job.get("date",""), "worker": new_job.get("worker","")}
        )

    return render_template('add_job.html',
                           workers=workers, machines=machines, locations=locations,
                           owners=owners, tenants=tenants, prev={}, job_registered=False)

@app.route('/edit_job/<int:job_index>', methods=['GET', 'POST'])
@perm_required('manage_jobs')
def edit_job(job_index):
    username = session['username']
    users_db = load_json('users.json', {})
    company  = users_db.get(username, {}).get('company')

    jobs_db  = load_json('jobs.json', {})
    job_list = jobs_db.get(company, [])
    if not (0 <= job_index < len(job_list)):
        return "작업을 찾을 수 없습니다.", 404
    job = job_list[job_index]

    machines  = load_json('machines.json',  {}).get(company, [])
    workers   = load_json('workers.json',   {}).get(company, [])
    locations = load_json('locations.json', {}).get(company, [])
    partners  = load_partners(company)
    owners    = partners.get('owners', [])
    tenants   = partners.get('tenants', [])

    if request.method == 'POST':
        t = (request.form.get('time') or '').strip()
        if not t:
            hh = (request.form.get('hour') or '').strip()
            mm = (request.form.get('minute') or '').strip()
            if hh and mm: t = f"{hh}:{mm}"
        job['time']           = t
        job['date']           = (request.form.get('date') or '').strip()
        job['worker']         = (request.form.get('worker') or '').strip()
        job['machine_name']   = (request.form.get('machine_name') or '').strip()
        job['machine_number'] = (request.form.get('machine_number') or '').strip()
        job['machine_alias']  = (request.form.get('machine_alias') or '').strip()

        client_primary = (request.form.get('client_primary') or '').strip()
        client_tenant  = (request.form.get('client_tenant')  or '').strip()
        if request.form.get('same_as_owner'):
            client_tenant = client_primary
        job['client_primary'] = client_primary
        job['client_tenant']  = client_tenant
        job['client']         = client_primary

        if request.form.get('save_owner') == '1' and client_primary and client_primary not in owners:
            owners.append(client_primary)
        if request.form.get('save_tenant') == '1' and client_tenant and client_tenant not in tenants:
            tenants.append(client_tenant)
        save_partners(company, owners, tenants)

        job['location'] = (request.form.get('location') or '').strip()
        job['note']     = (request.form.get('note') or '').strip()
        job['is_spare'] = bool(request.form.get('is_spare'))

        ot = (request.form.get('outsource_type') or 'none').strip()
        if ot not in ('none', 'received', 'given'):
            ot = 'none'
        job['outsource_type']    = ot
        job['outsource_partner'] = (request.form.get('outsource_partner') or '').strip()

        amount_raw = (request.form.get('amount_man') or '').strip()
        try: amount_man = int(amount_raw) if amount_raw != '' else 0
        except ValueError: amount_man = 0
        job['amount_man']   = amount_man
        job['share_amount'] = bool(request.form.get('share_amount'))

        old_paid = int(job.get('paid_amount_man') or 0)
        if amount_man <= 0:
            job['paid_amount_man'] = 0
            job['payment_status']  = '미설정'
        else:
            paid = max(0, min(old_paid, amount_man))
            job['paid_amount_man'] = paid
            if paid == 0:
                job['payment_status'] = '미납'
            elif paid >= amount_man:
                job['payment_status'] = '완납'
            else:
                job['payment_status'] = '부분'

        job['duration_type']  = (request.form.get('duration_type') or '하루').strip()
        job['duration_hours'] = (request.form.get('duration_hours') or '').strip() if job['duration_type'] == 'N시간' else ''

        save_json('jobs.json', jobs_db)

        params = {}
        for k, v in request.form.items():
            if k.startswith('filter_') and v != '':
                params[k[7:]] = v
        return redirect(url_for('view_jobs', **params))

    return render_template('edit_job.html',
                           job=job, job_index=job_index,
                           machines=machines, workers=workers, locations=locations,
                           owners=owners, tenants=tenants)

@app.route('/delete_job/<int:job_index>')
@perm_required('manage_jobs')
def delete_job(job_index):
    username = session['username']
    users_db = load_json('users.json', {})
    company = users_db.get(username, {}).get('company', '')

    jobs_db = load_json('jobs.json', {})
    job_list = jobs_db.get(company, [])
    if not (0 <= job_index < len(job_list)):
        return "작업을 찾을 수 없습니다.", 404

    job_list.pop(job_index)
    save_json('jobs.json', jobs_db)
    return redirect(url_for('view_jobs', **request.args))

@app.route('/bulk_action', methods=['POST'])
@perm_required('manage_jobs')
def bulk_action():
    action = request.form.get('action', '').strip()
    try:
        selected = [int(x) for x in request.form.getlist('selected_jobs')]
    except ValueError:
        selected = []

    users_db = load_json('users.json', {})
    user = users_db.get(session['username'], {})
    company = user.get('company', '')

    db = load_json('jobs.json', {})
    jobs = db.get(company, [])

    if action == 'complete':
        for idx in selected:
            if 0 <= idx < len(jobs):
                jobs[idx]['status'] = '완료'
    elif action == 'delete':
        for idx in sorted(set(selected), reverse=True):
            if 0 <= idx < len(jobs):
                jobs.pop(idx)

    db[company] = jobs
    save_json('jobs.json', db)
    return redirect(url_for('view_jobs'))

@app.route('/export_selected', methods=['POST'])
@perm_required('manage_jobs')
def export_selected():
    users_db = load_json('users.json', {})
    user = users_db.get(session['username'], {})
    company = user.get('company', '')

    try:
        selected = [int(x) for x in request.form.getlist('selected_jobs')]
    except ValueError:
        selected = []
    if not selected:
        return back_with_error("선택된 작업이 없습니다.")

    db = load_json('jobs.json', {})
    jobs = db.get(company, [])

    rows = []
    for idx in selected:
        if 0 <= idx < len(jobs):
            j = jobs[idx]
            rows.append([
                idx,
                j.get('date',''),
                j.get('time',''),
                j.get('worker',''),
                j.get('machine_name',''),
                j.get('machine_number',''),
                j.get('machine_alias',''),
                (j.get('client_primary') or j.get('client','')),
                j.get('client_tenant',''),
                j.get('location',''),
                j.get('status',''),
                j.get('duration_type',''),
                j.get('duration_hours',''),
                int(j.get('amount_man') or 0),
                'Y' if j.get('share_amount') else '',
                j.get('outsource_type',''),
                j.get('outsource_partner',''),
                j.get('payment_status',''),
                int(j.get('paid_amount_man') or 0),
            ])

    if not rows:
        return back_with_error("선택된 작업이 없습니다.")

    filename = f"jobs_selected_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    s = StringIO(); s.write('\ufeff')
    writer = csv.writer(s)
    writer.writerow([
        'index','date','time','worker',
        'machine_name','machine_number','machine_alias',
        'client_primary','client_tenant','location',
        'status','duration_type','duration_hours',
        'amount_man','share_amount',
        'outsource_type','outsource_partner',
        'payment_status','paid_amount_man'
    ])
    writer.writerows(rows)
    data = s.getvalue()

    return Response(
        data,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@app.route('/api/toggle_complete/<int:job_index>', methods=['POST'])
def toggle_complete_api(job_index):
    if 'username' not in session:
        return jsonify(success=False, error='unauthorized'), 401

    users_db = load_json('users.json', {})
    user = users_db.get(session['username'], {})
    company = user.get('company', '')
    role = (user.get('role') or '').strip()
    username = session['username']

    db = load_json('jobs.json', {})
    jobs = db.get(company, [])
    if not (0 <= job_index < len(jobs)):
        return jsonify(success=False, error='index_out_of_range'), 400

    if role == 'worker' and (jobs[job_index].get('worker') or '') != username:
        return jsonify(success=False, error='forbidden'), 403

    cur = (jobs[job_index].get('status') or '진행중').strip()
    new_status = '완료' if cur != '완료' else '진행중'
    jobs[job_index]['status'] = new_status

    db[company] = jobs
    save_json('jobs.json', db)
    return jsonify(success=True, status=new_status)

# 캘린더
@app.route('/calendar')
def calendar_view():
    if 'username' not in session:
        return redirect(url_for('login'))

    users_db = load_json('users.json', {})
    user = users_db.get(session['username'], {})
    company = user.get('company', '')
    jobs = load_json('jobs.json', {}).get(company, [])

    total_count = len(jobs)
    complete_count = sum(1 for j in jobs if (j.get('status') or '').strip() == '완료')
    pending_count = total_count - complete_count

    events = []
    for j in jobs:
        d = (j.get('date') or '').strip()
        if not d: continue
        title_bits = [j.get('worker') or '', j.get('client') or '']
        title = " - ".join([x for x in title_bits if x]) or "작업"
        events.append({
            "title": title,
            "start": d,
            "extendedProps": {
                "status": j.get('status',''),
                "time": j.get('time',''),
                "duration_type": j.get('duration_type','하루'),
                "duration_hours": j.get('duration_hours','')
            }
        })

    return render_template(
        'calendar_jobs.html',
        total_count=total_count,
        complete_count=complete_count,
        pending_count=pending_count,
        events=json.dumps(events, ensure_ascii=False)
    )

@app.route('/api/payment/<int:job_index>', methods=['POST'], endpoint='payment_api')
@perm_required('manage_payments')
def payment_api(job_index):
    users_db = load_json('users.json', {})
    user = users_db.get(session['username'], {})
    company = user.get('company', '')

    db = load_json('jobs.json', {})
    jobs = db.get(company, [])
    if not (0 <= job_index < len(jobs)):
        return jsonify(success=False, error='index_out_of_range'), 400

    job = jobs[job_index]
    amount = int(job.get('amount_man') or 0)
    paid   = int(job.get('paid_amount_man') or 0)

    data = request.get_json(silent=True) or {}
    action = (data.get('action') or '').strip()

    if action == 'full':
        if amount <= 0: return jsonify(success=False, error='no_amount'), 400
        paid = amount; status = '완납'
    elif action == 'unpay':
        paid = 0; status = '미납' if amount > 0 else '미설정'
    elif action == 'partial':
        if amount <= 0: return jsonify(success=False, error='no_amount'), 400
        try: add = int(data.get('amount_man'))
        except (TypeError, ValueError):
            return jsonify(success=False, error='invalid_amount'), 400
        paid = max(0, min(amount, paid + add))
        if paid == 0: status = '미납'
        elif paid >= amount: status = '완납'
        else: status = '부분'
    else:
        return jsonify(success=False, error='invalid_action'), 400

    job['paid_amount_man'] = paid
    job['payment_status']  = status
    db[company] = jobs
    save_json('jobs.json', db)

    remaining = max(0, amount - paid)
    return jsonify(success=True, payment_status=status, remaining=remaining,
                   paid_amount_man=paid, amount_man=amount)

# =========================================================
# 직원/장비/거래처/회사 (원기능 유지)
# =========================================================
@app.route('/add_worker', methods=['GET', 'POST'])
@perm_required('manage_workers')
def add_worker():
    users_db = load_json('users.json', {})
    company = users_db.get(session['username'], {}).get('company', '')

    workers_db = load_json('workers.json', {})
    workers = workers_db.get(company, [])

    if request.method == 'POST':
        name  = (request.form.get('name')  or '').strip()
        phone = (request.form.get('phone') or '').strip()

        if not name or not phone:
            return render_template('add_worker.html', workers=workers,
                                   error="기사 이름과 전화번호를 모두 입력해 주세요.")

        if any(w.get('name') == name or w.get('phone') == phone for w in workers):
            return render_template('add_worker.html', workers=workers,
                                   error="이미 등록된 기사명 또는 전화번호입니다.")

        username = f"{company}{name}"

        new_worker = {
            "username": username,
            "name": name,
            "phone": phone,
            "role": "worker",
            "status": "active"
        }
        workers.append(new_worker)
        workers_db[company] = workers
        save_json('workers.json', workers_db)

        udb = load_json('users.json', {})
        default_pw = (phone[-4:] if len(phone) >= 4 else phone) or "0000"
        udb[username] = {
            'password': udb.get(username, {}).get('password', default_pw),
            'role': 'worker',
            'company': company,
            'name': name,
            'phone': phone,
            'status': 'active'
        }
        save_json('users.json', udb)

        return redirect_with_from('add_worker')

    err = request.args.get('error')
    return render_template('add_worker.html', workers=workers, error=err)

@app.route('/manage_machines', methods=['GET','POST'], endpoint='manage_machines')
@app.route('/machines',        methods=['GET','POST'])
def manage_machines_alias():
    return add_machine()

@app.route('/add_machine', methods=['GET', 'POST'])
@perm_required('manage_machines')
def add_machine():
    users_db = load_json('users.json', {})
    company = users_db[session['username']]['company']

    machines_db = load_json('machines.json', {})
    machines = machines_db.get(company, [])
    error = None

    if request.method == 'POST':
        action = (request.form.get('action') or 'add').strip()

        if action == 'add':
            name = request.form.get('machine_name', '').strip()
            number = request.form.get('machine_number', '').strip()
            alias = request.form.get('machine_alias', '').strip()

            if not name or not number:
                error = "장비명과 차량번호는 필수 입력입니다."
            elif any(m.get('number') == number for m in machines):
                error = f"차량번호 {number} 는 이미 등록되어 있습니다."
            else:
                machines.append({'name': name, 'number': number, 'alias': alias})
                machines_db[company] = machines
                save_json('machines.json', machines_db)
                return redirect_with_from('add_machine')

        elif action == 'edit_save':
            try:
                idx = int(request.form.get('edit_idx', '-1'))
            except ValueError:
                idx = -1
            if 0 <= idx < len(machines):
                machines[idx]['name'] = request.form.get('machine_name', '').strip()
                machines[idx]['number'] = request.form.get('machine_number', '').strip()
                machines[idx]['alias'] = request.form.get('machine_alias', '').strip()
                machines_db[company] = machines
                save_json('machines.json', machines_db)
            return redirect_with_from('add_machine')

        elif action == 'delete':
            number = request.form.get('machine_number', '').strip()
            machines = [m for m in machines if m.get('number') != number]
            machines_db[company] = machines
            save_json('machines.json', machines_db)
            return redirect_with_from('add_machine')

    edit_machine = None
    edit_machine_idx = None
    edit_idx = request.args.get('edit')
    if edit_idx is not None and str(edit_idx).isdigit():
        idx = int(edit_idx)
        if 0 <= idx < len(machines):
            edit_machine = machines[idx]
            edit_machine_idx = idx

    return render_template(
        'add_machine.html',
        machines=machines,
        error=error,
        edit_machine=edit_machine,
        edit_machine_idx=edit_machine_idx
    )

@app.route('/manage_workers')
@perm_required('manage_workers')
def manage_workers():
    company = session['company']
    workers_db = load_json('workers.json', {})
    users_db = load_json('users.json', {})

    workers = workers_db.get(company, [])
    for w in workers:
        user_info = users_db.get(w['username'])
        w['role'] = (user_info or {}).get('role', 'worker')
    return render_template('manage_workers.html', workers=workers)

@app.route('/approve_worker/<username>', methods=['POST'])
@perm_required('approve_workers')
def approve_worker(username):
    users = load_json('users.json', {})
    workers_db = load_json('workers.json', {})
    company = load_json('users.json', {}).get(session['username'], {}).get('company','')

    u = users.get(username)
    if not u:
        return back_with_error("사용자 정보를 찾을 수 없습니다.")

    u['status'] = 'active'
    pending = (u.get('pending_update') or {})
    if 'phone' in pending and pending['phone']:
        u['phone'] = pending['phone']
        u.pop('pending_update', None)

    lst = workers_db.get(u.get('company','') or company, [])
    for w in lst:
        if w.get('username') == username:
            w['status'] = 'active'
            if 'phone' in pending and pending['phone']:
                w['phone'] = pending['phone']
            break
    workers_db[u.get('company','') or company] = lst

    save_json('users.json', users)
    save_json('workers.json', workers_db)
    return redirect_with_from('add_worker')

@app.route('/delete_worker', methods=['POST'])
@app.route('/delete_worker/<username>', methods=['POST'])
@perm_required('manage_workers')
def delete_worker(username=None):
    username = username or request.form.get('username') or request.args.get('username')
    if not username:
        return back_with_error('add_worker', '대상 사용자를 찾을 수 없습니다.')

    users = load_json('users.json', {})
    workers_db = load_json('workers.json', {})
    company = load_json('users.json', {}).get(session['username'], {}).get('company','')

    lst = workers_db.get(company, [])
    workers_db[company] = [w for w in lst if w.get('username') != username]
    if username in users:
        users.pop(username, None)

    save_json('workers.json', workers_db)
    save_json('users.json', users)
    return redirect_with_from('add_worker')

@app.route('/grant_manager', methods=['POST'])
@app.route('/grant_manager/<username>', methods=['POST'])
@perm_required('manage_workers')
def grant_manager(username=None):
    username = username or request.form.get('username') or request.args.get('username')
    if not username:
        return back_with_error('add_worker', '대상 사용자를 찾을 수 없습니다.')

    users = load_json('users.json', {})
    workers_db = load_json('workers.json', {})
    company = load_json('users.json', {}).get(session['username'], {}).get('company','')

    def ensure_user_entry(users_db, workers_db, company, username):
        if username in users_db:
            return users_db[username]
        worker = None
        for w in workers_db.get(company, []):
            if w.get('username') == username:
                worker = w; break
        if not worker:
            return None
        phone = (worker.get('phone') or '').strip()
        default_pw = (phone[-4:] if len(phone) >= 4 else phone) or "0000"
        users_db[username] = {
            'password': default_pw,
            'role': worker.get('role') or 'worker',
            'company': company,
            'name': worker.get('name', ''),
            'phone': phone,
            'status': worker.get('status') or 'active',
        }
        return users_db[username]

    u = ensure_user_entry(users, workers_db, company, username)
    if not u:
        return back_with_error('add_worker', '사용자 정보를 찾을 수 없습니다.')

    u['role'] = 'manager'
    for w in workers_db.get(company, []):
        if w.get('username') == username:
            w['role'] = 'manager'
            break

    save_json('users.json', users)
    save_json('workers.json', workers_db)
    return redirect_with_from('add_worker')

@app.route('/revoke_manager', methods=['POST'])
@app.route('/revoke_manager/<username>', methods=['POST'])
@perm_required('manage_workers')
def revoke_manager(username=None):
    username = username or request.form.get('username') or request.args.get('username')
    if not username:
        return back_with_error('add_worker', '대상 사용자를 찾을 수 없습니다.')

    users = load_json('users.json', {})
    workers_db = load_json('workers.json', {})
    company = load_json('users.json', {}).get(session['username'], {}).get('company','')

    if username not in users:
        return back_with_error('add_worker', '사용자 정보를 찾을 수 없습니다.')
    if users[username].get('role') == 'boss':
        return back_with_error('add_worker', '사장 권한은 해제할 수 없습니다.')

    users[username]['role'] = 'worker'
    for w in workers_db.get(company, []):
        if w.get('username') == username:
            w['role'] = 'worker'
            break

    save_json('users.json', users)
    save_json('workers.json', workers_db)
    return redirect_with_from('add_worker')

@app.route('/update_worker/<username>', methods=['GET', 'POST'])
@perm_required('manage_workers')
def update_worker(username):
    company = session['company']
    users_db = load_json('users.json', {})
    workers_db = load_json('workers.json', {})

    user = users_db.get(username)
    if not user or user.get('company') != company:
        return redirect('/manage_workers')

    if user.get('role') == 'boss':
        return "보스 계정은 수정할 수 없습니다.", 403

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        if name:  user['name'] = name
        if phone: user['phone'] = phone
        save_json('users.json', users_db)

        for w in workers_db.get(company, []):
            if w.get('username') == username:
                if name:  w['name'] = name
                if phone: w['phone'] = phone
                break
        workers_db[company] = workers_db.get(company, [])
        save_json('workers.json', workers_db)
        return redirect('/manage_workers')

    return render_template('update_worker.html', user=user)

@app.route('/clients', methods=['GET', 'POST'])
@perm_required('manage_clients')
def manage_clients():
    user = get_current_user()
    company = user.get('company', '')
    entry   = load_partners(company)
    owners  = entry.get('owners', [])
    tenants = entry.get('tenants', [])

    if request.method == 'POST':
        kind   = (request.form.get('kind')   or '').strip()    # 'primary' | 'tenant'
        action = (request.form.get('action') or '').strip()    # 'add' | 'delete'
        name   = (request.form.get('name')   or '').strip()

        if kind not in ('primary', 'tenant') or action not in ('add', 'delete'):
            return redirect_with_from('manage_clients')

        lst = owners if kind == 'primary' else tenants
        if action == 'add':
            if name and name not in lst:
                lst.append(name)
        elif action == 'delete':
            idx = request.form.get('idx')
            if idx not in (None, ''):
                try:
                    i = int(idx)
                    if 0 <= i < len(lst):
                        lst.pop(i)
                except ValueError:
                    pass
            elif name in lst:
                lst.remove(name)

        save_partners(company, owners, tenants)
        return redirect_with_from('manage_clients')

    return render_template('manage_clients.html', owners=owners, tenants=tenants)

@app.route('/company_info', methods=['GET', 'POST'])
@perm_required('manage_roles')
def company_info():
    username = session['username']
    users_db = load_json('users.json', {})
    user = users_db.get(username, {})

    if user.get('role') != 'boss':
        return "권한이 없습니다.", 403

    company = user.get('company')
    companies = load_json('companies.json', {})
    company_info = companies.get(company, {})

    error = None
    success = None

    if request.method == 'POST':
        new_company_name = request.form['company']
        new_phone = request.form['phone']
        new_password = request.form['password']
        new_company_code = request.form['company_code']

        if len(new_company_code) != 6:
            error = '회사 코드는 6자리여야 합니다.'
        elif new_company_name != company and new_company_name in companies:
            error = '이미 존재하는 회사명입니다.'
        else:
            if new_company_name != company:
                moved = companies.pop(company, {})
                companies[new_company_name] = moved
                users_db[username]['company'] = new_company_name
                company = new_company_name

            users_db[username]['phone'] = new_phone
            if new_password.strip():
                users_db[username]['password'] = new_password
            save_json('users.json', users_db)

            companies.setdefault(company, {})
            companies[company]['phone'] = new_phone
            companies[company]['code'] = new_company_code
            save_json('companies.json', companies)

            success = '회사 정보가 성공적으로 수정되었습니다.'

    return render_template(
        'company_info.html',
        username=username,
        user=user,
        company=company,
        company_info=company_info,
        error=error,
        success=success
    )

# 회원가입 (보스/기사)
@app.route('/register/boss', methods=['GET', 'POST'])
def register_boss():
    if request.method == 'POST':
        password     = (request.form.get('password') or '').strip()
        company      = (request.form.get('company') or '').strip()
        phone        = (request.form.get('phone') or '').strip()
        company_code = (request.form.get('company_code') or '').strip()

        if not (password and company and phone and len(company_code) == 6):
            return render_template('register_boss.html', error='모든 값을 올바르게 입력해 주세요.')

        users = load_json('users.json', {})
        companies = load_json('companies.json', {})

        if company in companies:
            return render_template('register_boss.html', error='이미 존재하는 회사명입니다.')
        if any(u.get('phone') == phone for u in users.values()):
            return render_template('register_boss.html', error='해당 전화번호로 이미 가입된 계정이 있습니다.')

        base_username = f"{company}boss"
        username = base_username
        i = 1
        while username in users:
            username = f"{base_username}{i}"
            i += 1

        users[username] = {
            "password": password,
            "role": "boss",
            "company": company,
            "phone": phone,
            "company_code": company_code,
            "name": "사장님",
        }
        save_json('users.json', users)

        companies[company] = {"code": company_code, "phone": phone}
        save_json('companies.json', companies)

        _seed_company_containers(company)

        return render_template(
            'register_boss_success.html',
            boss_name="사장님",
            company=company,
            phone=phone,
            company_code=company_code
        )

    return render_template('register_boss.html')

@app.route('/_repair_companies_once')
def _repair_companies_once():
    users = load_json('users.json', {})
    companies = load_json('companies.json', {})
    changed = 0
    for u in users.values():
        if u.get('role') == 'boss':
            c = (u.get('company') or '').strip()
            if not c: continue
            code = (u.get('company_code') or '').strip()
            phone = (u.get('phone') or '').strip()
            if c not in companies or companies[c].get('code') != code or companies[c].get('phone') != phone:
                companies[c] = {"code": code, "phone": phone}
                changed += 1
    save_json('companies.json', companies)
    return f"repaired: {changed}"

@app.route('/register/worker', methods=['GET', 'POST'])
def register_worker():
    try:
        companies = load_json('companies.json', {})

        if request.method == 'POST':
            name = (request.form.get('name') or '').strip()
            phone = (request.form.get('phone') or '').strip()
            company = (request.form.get('company') or '').strip()
            input_code = (request.form.get('company_code') or '').strip()
            password = (request.form.get('password') or '').strip()

            if company not in companies:
                error = '존재하지 않는 회사명입니다.'
                return render_template('register_worker.html', companies=sorted(companies.keys()), error=error)

            if companies[company]['code'] != input_code:
                error = '회사 코드가 올바르지 않습니다.'
                return render_template('register_worker.html', companies=sorted(companies.keys()), error=error)

            users_db = load_json('users.json', {})
            workers_db = load_json('workers.json', {})
            workers_db.setdefault(company, [])

            for u_name, u in users_db.items():
                if u.get('company') == company and (u.get('phone') or '') == phone:
                    return render_template(
                        'register_worker_conflict.html',
                        company=company,
                        name=name,
                        existing_name=u.get('name',''),
                        existing_phone=phone,
                        message='이미 동일한 전화번호로 가입된 계정이 있습니다.',
                        show_homonym=False
                    )

            exists = None
            for w in workers_db[company]:
                if (w.get('name') or '').strip() == name:
                    exists = w
                    break

            if exists:
                ex_username = exists.get('username') or f"{company}{name}"

                def ensure_user_entry(users_db, workers_db, company, username):
                    if username in users_db:
                        return users_db[username]
                    worker = None
                    for w in workers_db.get(company, []):
                        if w.get('username') == username:
                            worker = w; break
                    if not worker:
                        return None
                    phone = (worker.get('phone') or '').strip()
                    default_pw = (phone[-4:] if len(phone) >= 4 else phone) or "0000"
                    users_db[username] = {
                        'password': default_pw,
                        'role': worker.get('role') or 'worker',
                        'company': company,
                        'name': worker.get('name', ''),
                        'phone': phone,
                        'status': worker.get('status') or 'active',
                    }
                    return users_db[username]

                ensure_user_entry(users_db, workers_db, company, ex_username)
                users_db[ex_username]['status'] = 'pending'
                users_db[ex_username]['pending_update'] = {'phone': phone}

                for w in workers_db[company]:
                    if w.get('username') == ex_username:
                        w['status'] = 'pending'
                        break

                save_json('users.json', users_db)
                save_json('workers.json', workers_db)

                return render_template(
                    'register_worker_conflict.html',
                    company=company,
                    name=name,
                    existing_name=exists.get('name',''),
                    existing_phone=exists.get('phone',''),
                    new_phone=phone,
                    password=password,
                    message=f'해당 회사에 {name}({exists.get("phone","")}) 기사가 존재합니다. 관리자/사장님에게 가입승인을 요청했습니다.',
                    show_homonym=True
                )

            base_username = f"{company}{name}"
            username = base_username
            suffix = 1
            while username in users_db:
                username = f"{base_username}{suffix}"
                suffix += 1

            users_db[username] = {
                'password': password,
                'role': 'worker',
                'company': company,
                'name': name,
                'phone': phone,
                'status': 'pending'
            }
            save_json('users.json', users_db)

            workers_db[company] = [w for w in workers_db[company] if w.get('phone') != phone]
            workers_db[company].append({
                'username': username,
                'name': name,
                'phone': phone,
                'role': 'worker',
                'status': 'pending'
            })
            save_json('workers.json', workers_db)

            return render_template('register_worker_pending.html', name=name, company=company)

        return render_template('register_worker.html', companies=sorted(companies.keys()))

    except Exception as e:
        import traceback
        return f"<h2>서버 오류 발생:<br>{e}</h2><pre>{traceback.format_exc()}</pre>"

@app.route('/register/worker/resolve_homonym', methods=['POST'])
def resolve_homonym():
    name = (request.form.get('name') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    company = (request.form.get('company') or '').strip()
    password = (request.form.get('password') or '').strip()

    users_db = load_json('users.json', {})
    workers_db = load_json('workers.json', {})
    workers_db.setdefault(company, [])

    ex = None
    for w in workers_db[company]:
        if (w.get('name') or '').strip() == name and (w.get('status') or '') == 'pending':
            ex = w; break
    if not ex:
        for w in workers_db[company]:
            if (w.get('name') or '').strip() == name:
                ex = w; break
    if not ex:
        return back_with_error("기존 기사를 찾을 수 없습니다.")

    ex_username = ex.get('username') or f"{company}{name}"

    ex['name'] = f"{name}(A)"
    ex['status'] = 'active'
    if ex_username in users_db:
        users_db[ex_username]['name'] = f"{name}(A)"
        users_db[ex_username]['status'] = 'active'
        users_db[ex_username].pop('pending_update', None)

    base_username = f"{company}{name}"
    username = base_username
    i = 1
    while username in users_db:
        username = f"{base_username}{i}"
        i += 1

    users_db[username] = {
        'password': password,
        'role': 'worker',
        'company': company,
        'name': f"{name}(B)", 
        'phone': phone,
        'status': 'pending'
    }
    workers_db[company].append({
        'username': username,
        'name': f"{name}(B)",
        'phone': phone,
        'role': 'worker',
        'status': 'pending'
    })

    save_json('users.json', users_db)
    save_json('workers.json', workers_db)

    return render_template('register_worker_pending.html', name=f"{name}(B)", company=company)

@app.route('/edit_worker', methods=['GET', 'POST'])
def edit_worker():
    if 'username' not in session:
        return redirect('/login')

    worker_username = request.args.get('worker_username')
    users_db = load_json('users.json', {})
    current_user = users_db.get(session['username'], {})
    is_admin_user = current_user.get('role') in ('boss', 'manager')

    if worker_username and is_admin_user:
        username_to_edit = worker_username
        back_endpoint = 'add_worker'
        back_label = '← 기사 등록/관리로 돌아가기'
    else:
        username_to_edit = session['username']
        back_endpoint = 'dashboard_worker' if current_user.get('role') == 'worker' else 'dashboard'
        back_label = '🏠 메인화면으로 돌아가기'

    user_info = users_db.get(username_to_edit)
    if not user_info:
        return "사용자 정보를 찾을 수 없습니다.", 404

    if user_info.get('role') == 'boss' and not (username_to_edit == session['username']):
        return "보스 계정은 수정할 수 없습니다.", 403

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        password = (request.form.get('password') or '').strip()

        if name:     user_info['name'] = name
        if phone:    user_info['phone'] = phone
        if password: user_info['password'] = password

        users_db[username_to_edit] = user_info
        save_json('users.json', users_db)

        target_company = user_info.get('company', '')
        workers_db = load_json('workers.json', {})
        lst = workers_db.get(target_company, [])
        for w in lst:
            if w.get('username') == username_to_edit:
                if name:  w['name'] = name
                if phone: w['phone'] = phone
                break
        workers_db[target_company] = lst
        save_json('workers.json', workers_db)

        return redirect_with_from(back_endpoint)

    return render_template('edit_worker.html', user=user_info, back_endpoint=back_endpoint, back_label=back_label)

# =========================================================
# 재무(요약/수입/지출)
# =========================================================
@app.route("/finance")
def finance_dashboard():
    company = (session.get("company") or (get_current_user() or {}).get("company", "")).strip()

    def truthy(v): return str(v or "").lower() in ("1","true","y","yes","on")
    def norm_pay(v):
        v = (v or "").strip().lower()
        if v in ("unpaid","dues","미납","미정산","미납만","미정산만"): return "unpaid"
        if v in ("paid","완납","완납만"): return "paid"
        return ""
    def norm_status(v):
        v = (v or "").strip().lower()
        if v in ("todo","progress","ongoing","진행","진행중"): return "todo"
        if v in ("done","completed","완료"): return "done"
        return ""

    today = date.today()
    ytd_start = date(today.year, 1, 1)
    start = _parse_date_safe(request.args.get("start") or ytd_start.isoformat()) or ytd_start
    end   = _parse_date_safe(request.args.get("end")   or today.isoformat())      or today
    ytd_start_str, today_str = ytd_start.isoformat(), today.isoformat()

    tab  = (request.args.get("tab") or "summary").strip()
    half = (request.args.get("half") or "").strip()

    by_worker  = truthy(request.args.get("by_worker", "0"))
    by_machine = truthy(request.args.get("by_machine", "0"))
    by_client  = truthy(request.args.get("by_client", "0"))

    pay      = norm_pay(request.args.get("pay"))
    status_f = norm_status(request.args.get("status"))
    unpaid_only = truthy(request.args.get("unpaid_only"))

    sel_worker   = (request.args.get("worker") or "").strip()
    input_worker = (request.args.get("worker_input") or "").strip()
    q_plate      = (request.args.get("plate") or "").strip()
    q_owner      = (request.args.get("owner") or "").strip()
    q_tenant     = (request.args.get("tenant") or "").strip()

    inc_cat  = (request.args.get("inc_cat")  or "").strip()
    inc_desc = (request.args.get("inc_desc") or "").strip()
    exp_cat  = (request.args.get("exp_cat")  or "").strip()
    exp_desc = (request.args.get("exp_desc") or "").strip()

    try:
        page_size = int(request.args.get("page_size", "20"))
    except Exception:
        page_size = 20
    if page_size not in (20, 100):
        page_size = 20
    try:
        page = max(1, int(request.args.get("page", "1")))
    except Exception:
        page = 1

    def _half_ok(d: date) -> bool:
        if not d: return False
        if half == "H12": return 1 <= d.month <= 6
        if half == "H34": return 7 <= d.month <= 12
        return True

    # ===== 작업 로드 + 외주 자동동기화 =====
    jobs_db = load_json("jobs.json", {})
    jobs_all_company = jobs_db.get(company, [])
    for i, j in enumerate(jobs_all_company):
        j["_idx"] = i
    _sync_outsourcing_entries(company, jobs_all_company)  # 버튼 없이 자동

    incomes_db  = load_json("incomes.json", {})
    expenses_db = load_json("expenses_db.json", {})
    inc_all = incomes_db.get(company, [])
    exp_all = expenses_db.get(company, [])

    # 기간 필터
    inc_in_range = [r for r in inc_all if _in_range(_parse_date_safe(r.get("date")), start, end)]
    exp_in_range = [r for r in exp_all if _in_range(_parse_date_safe(r.get("date")), start, end)]

    # 검색 필터
    if inc_cat:
        inc_in_range = [r for r in inc_in_range if inc_cat.lower() in (r.get("category","").lower())]
    if inc_desc:
        inc_in_range = [r for r in inc_in_range if inc_desc.lower() in (r.get("desc","").lower())]
    if exp_cat:
        exp_in_range = [r for r in exp_in_range if exp_cat.lower() in (r.get("category","").lower())]
    if exp_desc:
        exp_in_range = [r for r in exp_in_range if exp_desc.lower() in (r.get("desc","").lower())]

    # ===== 외주 자동항목 ↔ 원본 작업 매핑 =====
    import hashlib as _h
    def _out_key_for_job(j: dict, kind: str) -> str:
        base = "|".join([
            str(j.get("date","")),
            str(j.get("worker","")),
            str(j.get("machine_number","")),
            str(j.get("outsource_partner","")),
            kind.lower().strip()
        ])
        return "auto-" + _h.md5(base.encode("utf-8")).hexdigest()[:16]

    job_info_by_key = {}
    for j in jobs_all_company:
        ot = (j.get("outsource_type") or "").lower()
        if ot in ("received","given"):
            k = _out_key_for_job(j, ot)
            job_info_by_key[k] = {
                "status": (j.get("status") or "").strip(),
                "payment_status": (j.get("payment_status") or "").strip(),
                "date": j.get("date") or "",
                "worker": (j.get("worker") or j.get("driver") or ""),
            }

    # ===== 표시 텍스트/링크 주입 =====
    # (1) 수입: 외주받음이 '완납'이 아니면 금액 칸에 텍스트, 외주 자동항목에만 링크
    for i in inc_in_range:
        if (i.get("source") or "") == "auto_outsrc_received":
            k  = i.get("auto_key") or i.get("id")
            ji = job_info_by_key.get(k, {})
            if ji:
                i["detail_url"] = url_for("view_jobs", date=ji.get("date",""), worker=ji.get("worker",""))
                if ji.get("payment_status") != "완납":
                    i["display_amount_text"] = "정산되지않음"

    # (2) 지출: 외주줬음은 표시금액 0, 진행중이면 텍스트, 외주 자동항목에만 링크
    def _is_auto_given(e):
        return (e.get("source") == "auto_outsrc_given") or (
            e.get("category") == "외주줬음" and e.get("auto_key")
        )
    for e in exp_in_range:
        if _is_auto_given(e):
            k  = e.get("auto_key") or e.get("id")
            ji = job_info_by_key.get(k, {})
            if ji:
                e["detail_url"] = url_for("view_jobs", date=ji.get("date",""), worker=ji.get("worker",""))
            if ji and ji.get("status") != "완료":
                e["display_amount_text"] = "작업 진행중"
        amt = _to_number(e.get("amount"))
        e["display_amount"] = 0 if _is_auto_given(e) else amt

    # 합계: 수입은 외주받음 중 '완납'만 포함
    def _count_income(i: dict) -> bool:
        if (i.get("source") or "") != "auto_outsrc_received":
            return True
        k  = i.get("auto_key") or i.get("id")
        ji = job_info_by_key.get(k)
        return bool(ji and ji.get("payment_status") == "완납")

    income_total  = int(sum(_to_number(i.get("amount")) for i in inc_in_range if _count_income(i)))
    expense_total = int(sum(e.get("display_amount", _to_number(e.get("amount"))) for e in exp_in_range))

    # ===== 상단 매출/미수 (외주받음 제외) =====
    import datetime as _dt
    completed, outstanding_all = [], 0
    for j in jobs_all_company:
        d = _parse_date_safe(j.get("date"))
        if not (_in_range(d, start, end) and _half_ok(d)):
            continue
        amt  = _amount_won(j)
        paid = _paid_won(j)
        outstanding_all += max(0, amt - min(amt, paid))
        if (j.get("status") or "진행중").strip() == "완료":
            if (j.get("outsource_type") or "none") != "received":
                completed.append((amt, paid))
    sales_total = sum(a for a, p in completed if a > 0 and p >= a)
    profit_total = int((sales_total or 0) - (expense_total or 0) + (income_total or 0))

    # ===== 요약/내역 표(외주 작업 제외) =====
    def _dt_key(j):
        d = (j.get("date") or "").strip()
        t = (j.get("time") or "").strip() or "00:00"
        try:
            return _dt.datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
        except:
            try:
                return _dt.datetime.strptime(d, "%Y-%m-%d")
            except:
                return _dt.datetime.min

    rows = []
    for j in jobs_all_company:
        d = _parse_date_safe(j.get("date"))
        if not (_in_range(d, start, end) and _half_ok(d)):
            continue
        if (j.get("outsource_type") or "none") in ("received", "given"):
            continue

        if by_worker:
            target = (sel_worker or input_worker).strip()
            if target and target not in ((j.get("worker") or j.get("driver") or "")):
                continue
        if by_machine and q_plate:
            plate = (j.get("machine_number") or j.get("plate") or "")
            if q_plate not in plate:
                continue
        if by_client:
            owner  = (j.get("client_primary") or j.get("client") or "")
            tenant = (j.get("client_tenant") or "")
            if q_owner and q_owner not in owner:    continue
            if q_tenant and q_tenant not in tenant: continue

        amt   = _amount_won(j)
        paid  = _paid_won(j)
        color = _color_by_payment(amt, paid, j.get("payment_status") or "")

        if unpaid_only or pay == "unpaid":
            if color not in ("unpaid","partial"):
                continue
        elif pay == "paid":
            if color != "paid":
                continue

        st = (j.get("status") or "진행중").strip()
        is_todo = (st != "완료")
        if status_f == "todo" and not is_todo:
            continue
        if status_f == "done" and is_todo:
            continue

        rows.append({
            "raw": j,
            "date": d or start,
            "date_str": j.get("date") or "",
            "owner": (j.get("client_primary") or j.get("client") or ""),
            "tenant": j.get("client_tenant") or "",
            "machine_name": (j.get("machine_name") or j.get("machine_alias") or j.get("machine") or ""),
            "machine_number": (j.get("machine_number") or j.get("plate") or ""),
            "worker": (j.get("worker") or j.get("driver") or j.get("기사") or "-"),
            "amount_won": amt,
            "paid_won": paid,
            "color": color,
            "is_todo": is_todo,
            "_sort": (_dt_key(j), j.get("_idx",-1)),
        })

    rows.sort(key=lambda x: x["_sort"], reverse=True)
    list_total_all = sum(r["amount_won"] for r in rows)
    total = len(rows)
    total_pages = max(1, (total + page_size - 1)//page_size)
    page = min(page, total_pages)
    sidx, eidx = (page - 1) * page_size, (page - 1) * page_size + page_size
    rows_page = rows[sidx:eidx]

    # 기사 드롭다운
    registered_workers = []
    workers_db = load_json("workers.json", {})
    for w in workers_db.get(company, []):
        nm = (w.get("name") or w.get("username") or "").strip()
        if nm and nm not in registered_workers:
            registered_workers.append(nm)
    for j in jobs_all_company:
        nm = (j.get("worker") or j.get("driver") or "").strip()
        if nm and nm not in registered_workers:
            registered_workers.append(nm)

    # 삭제 후에도 현재 필터 유지
    from urllib.parse import urlencode
    qs_common = {"start": start.isoformat(), "end": end.isoformat(), "page_size": page_size}
    qs_inc = urlencode(dict(qs_common, **{
        "tab": "income_list", "inc_cat": inc_cat, "inc_desc": inc_desc, "page": request.args.get("page","1"),
    }))
    qs_exp = urlencode(dict(qs_common, **{
        "tab": "expense_list","exp_cat": exp_cat,"exp_desc": exp_desc,"page": request.args.get("page","1"),
    }))

    # 수입/지출 정렬: (일자, 생성시각) 최신 우선
    def _rec_sort_key(r):
        d = _parse_date_safe(r.get("date")) or date.min
        ca = (r.get("created_at") or "")
        return (d, ca)
    inc_in_range.sort(key=_rec_sort_key, reverse=True)
    exp_in_range.sort(key=_rec_sort_key, reverse=True)

    return render_template(
        "finance.html",
        # 기간
        start=start.isoformat(), end=end.isoformat(),
        ytd_start=ytd_start_str, today_str=today_str,
        # 상단 지표
        sales_total=sales_total,
        sales_total_all=sales_total,
        unpaid_total=outstanding_all,
        expense_total=expense_total,
        income_total=income_total,
        profit_total=profit_total,
        # 목록(요약)
        jobs=rows_page, total=total, page=page, total_pages=total_pages, page_size=page_size,
        list_total_all=list_total_all,
        # 컨트롤 상태
        tab=tab, company=company, half=half,
        by_worker=by_worker, by_machine=by_machine, by_client=by_client,
        status=status_f, pay=pay, unpaid_only=unpaid_only,
        sel_worker=sel_worker, input_worker=input_worker, q_plate=q_plate,
        owner=q_owner, tenant=q_tenant,
        registered_workers=registered_workers,
        # 수입/지출 리스트 + 링크/표시텍스트
        incs=inc_in_range, inc_cat=inc_cat, inc_desc=inc_desc,
        exps=exp_in_range, exp_cat=exp_cat, exp_desc=exp_desc,
        # 삭제 후 필터 유지용 쿼리
        qs_inc=qs_inc, qs_exp=qs_exp,
    )

@app.post("/finance/job_delete")
def finance_delete_job():
    if 'username' not in session:
        return redirect(url_for('login'))

    user = get_current_user() or {}
    company = (user.get('company') or '').strip()

    try:
        job_index = int(request.form.get('job_index', '-1'))
    except Exception:
        job_index = -1

    db = load_json('jobs.json', {})
    arr = db.get(company, [])

    if 0 <= job_index < len(arr):
        del arr[job_index]
        db[company] = arr
        save_json('jobs.json', db)
        flash('작업이 삭제되었습니다.')
    else:
        flash('삭제 대상이 올바르지 않습니다.', 'error')

    params = _carry_params_from_form()
    if not params.get("tab"):
        params["tab"] = "summary"
    return redirect(url_for('finance_dashboard', **params))

@app.route("/finance/sync_outsourcing", methods=["GET", "POST"])
def finance_sync_outsourcing():
    if 'username' not in session:
        # fetch에서 HTML 리다이렉트를 받으면 파싱 실패하므로 200 + JSON으로 돌려줍니다.
        return jsonify(success=False, error="unauthorized"), 200

    try:
        user = get_current_user() or {}
        company = (user.get('company') or '').strip()

        # jobs 읽기
        jobs_db = load_json('jobs.json', {})
        jobs = (jobs_db.get(company, []) if isinstance(jobs_db, dict) else [])

        # incomes/expenses 내부에서 list/dict 혼합 상태가 들어와도 터지지 않게 보장
        # (_sync_outsourcing_entries) 진입 전에 파일 형태부터 맞춰둡니다.
        incomes_db  = _ensure_company_bucket(load_json("incomes.json", {}), company)
        expenses_db = _ensure_company_bucket(load_json("expenses_db.json", {}), company)
        save_json("incomes.json", incomes_db)
        save_json("expenses_db.json", expenses_db)

        summary = _sync_outsourcing_entries(company, jobs)
        return jsonify(success=True, **summary), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        # 500 대신 200으로 내려서 프론트 alert에 에러 메시지가 찍히도록
        return jsonify(success=False, error=str(e)), 200

# ----- 수입
@app.post("/finance/income/add")
def finance_income_add():
    if 'username' not in session:
        return redirect(url_for('login'))
    company = (get_current_user() or {}).get('company','')

    date_s = (request.form.get('date') or '').strip()
    cat  = (request.form.get('category') or '').strip()
    desc = (request.form.get('desc') or '').strip()
    amt  = int(_to_number(request.form.get('amount') or 0))  # _as_int 제거

    incomes_db = load_json("incomes.json", {})
    rows = list(incomes_db.get(company, []))
    rows.append({
        "id": uuid.uuid4().hex,
        "date": date_s,
        "category": cat,
        "desc": desc,
        "amount": amt,
        "created_at": _now_str(),
    })
    incomes_db[company] = rows
    save_json("incomes.json", incomes_db)

    params = _carry_params_from_form()
    params["tab"] = "income_list"
    return redirect(url_for('finance_dashboard', **params))

@app.post("/finance/income/<iid>/delete")
def finance_income_delete(iid):
    if 'username' not in session:
        return redirect(url_for('login'))
    company = (get_current_user() or {}).get('company','')

    incomes_db = load_json("incomes.json", {})
    rows = list(incomes_db.get(company, []))

    # 삭제 대상 찾기 (먼저 찾아야 원본 작업 삭제 가능)
    target = None
    for r in rows:
        if str(r.get("id")) == str(iid):
            target = r
            break

    # 자동 외주 수입이면 원본 작업도 제거
    if target and str(target.get("source","")).startswith("auto_outsrc_"):
        _delete_outsourcing_job_by_auto_item(company, target)

    # 수입 행 삭제
    rows = [r for r in rows if str(r.get("id")) != str(iid)]
    incomes_db[company] = rows
    save_json("incomes.json", incomes_db)

    # 🔁 쿼리스트링 유지해서 리다이렉트 (페이지/검색 유지)
    params = request.args.to_dict(flat=True)
    params["tab"] = "income_list"
    if "page" not in params:
        params["page"] = "1"
    return redirect(url_for('finance_dashboard', **params))


@app.post("/finance/income/export", endpoint="finance_income_export_xlsx")
def finance_income_export_xlsx():
    if 'username' not in session:
        return redirect(url_for('login'))

    from io import BytesIO
    from urllib.parse import quote
    import datetime as dt
    import hashlib as _h
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    user = get_current_user() or {}
    company = (user.get('company') or '').strip()

    # 요청 파라미터
    start = (request.form.get('start') or '').strip()
    end   = (request.form.get('end') or '').strip()
    inc_cat  = (request.form.get('inc_cat') or '').strip().lower()
    inc_desc = (request.form.get('inc_desc') or '').strip().lower()

    # 열 너비 자동조절
    def _autofit_worksheet(ws, min_widths=None, max_width=60):
        min_widths = min_widths or {}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                s = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, max((len(x) for x in s.splitlines()), default=0))
            ws.column_dimensions[letter].width = max(min_widths.get(col_idx, 10),
                                                     min(max_width, max_len + 2))

    # --- 외주받음 정산 상태 확인용 매핑 (jobs.json) ---
    def _out_key_for_job(j: dict, kind: str) -> str:
        base = "|".join([
            str(j.get("date","")),
            str(j.get("worker","")),
            str(j.get("machine_number","")),
            str(j.get("outsource_partner","")),
            kind.lower().strip()
        ])
        return "auto-" + _h.md5(base.encode("utf-8")).hexdigest()[:16]

    jobs = load_json("jobs.json", {}).get(company, [])
    job_pay_by_key = {}
    for j in jobs:
        if (j.get("outsource_type") or "").lower() == "received":
            k = _out_key_for_job(j, "received")
            job_pay_by_key[k] = (j.get("payment_status") or "").strip()

    # --- 데이터 필터링 (정산되지 않은 외주수입 제외) ---
    rows = []
    for r in incomes_read(company):
        d = (r.get('date') or '').strip()
        if not ((not start or d >= start) and (not end or d <= end)):
            continue
        if inc_cat and inc_cat not in (r.get('category') or '').lower():
            continue
        if inc_desc and inc_desc not in (r.get('desc') or '').lower():
            continue

        src = (r.get('source') or '')
        if src == "auto_outsrc_received":
            k = r.get("auto_key") or r.get("id")
            # 완납이 아닌 외주수입 → 제외
            if job_pay_by_key.get(k) != "완납":
                continue

        rows.append(r)

    # --- 엑셀 작성 ---
    wb = Workbook(); ws = wb.active; ws.title = "수입"
    headers = ["일자","카테고리","내역","금액(원)"]; ws.append(headers)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = center; c.border = thin

    total_income = 0
    for r in rows:
        amt = int(_to_number(r.get('amount')))
        total_income += amt
        row = [r.get('date',''), r.get('category',''), r.get('desc',''), amt]
        ws.append(row)

    # 본문 테두리/정렬/숫자서식
    if ws.max_row >= 2:
        amt_col = 4
        for rr in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for i, c in enumerate(rr, start=1):
                c.border = thin
                if i == amt_col:
                    c.number_format = "#,##0"
                    c.alignment = right
                else:
                    c.alignment = left

    # 합계 행
    ws.append(["총 수입","","", total_income])
    for i, c in enumerate(ws[ws.max_row], start=1):
        c.border = thin
        c.font = Font(bold=True)
        c.alignment = right if i == 4 else right
        if i == 4: c.number_format = "#,##0"

    # 열너비 자동
    _autofit_worksheet(ws, min_widths={1:12, 2:16, 3:24, 4:14})

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    from urllib.parse import quote
    start_dt   = _parse_date_safe(request.form.get('start')) or date.today().replace(month=1, day=1)
    end_dt     = _parse_date_safe(request.form.get('end'))   or date.today()
    start_code = start_dt.strftime('%Y%m%d')
    end_code   = end_dt.strftime('%Y%m%d')
    filename   = f"수입내역 {start_code}_{end_code}.xlsx"
    q = quote(filename)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={q}; filename*=UTF-8''{q}"}
    )

# ----- 지출
@app.route("/finance/expense/add", methods=["POST"])
def finance_expense_add():
    if 'username' not in session:
        return redirect(url_for('login'))
    company = (get_current_user() or {}).get("company","")

    d_str = (request.form.get("date") or date.today().isoformat()).strip()
    category = (request.form.get("category") or "").strip()
    desc = (request.form.get("desc") or "").strip()
    amount_raw = (request.form.get("amount") or "0").replace(",", "").strip()
    try:
        amount = float(amount_raw)
    except Exception:
        flash("지출 금액 형식이 올바르지 않습니다.")
        params = _carry_params_from_form(); params["tab"] = "expense_list"
        return redirect(url_for("finance_dashboard", **params))

    expenses_db = load_json("expenses_db.json", {})
    lst = list(expenses_db.get(company, []))
    lst.append({
        "id": uuid.uuid4().hex[:8],
        "date": d_str,
        "category": category,
        "desc": desc,
        "amount": amount,
        "created_at": _now_str(),
    })
    expenses_db[company] = lst
    save_json("expenses_db.json", expenses_db)
    flash("지출이 등록되었습니다.")

    params = _carry_params_from_form()
    params["tab"] = "expense_list"
    return redirect(url_for("finance_dashboard", **params))

@app.route("/finance/expense/delete/<eid>", methods=["POST"])
def finance_expense_delete(eid):
    if 'username' not in session:
        return redirect(url_for('login'))
    company = (get_current_user() or {}).get("company","")

    expenses_db = load_json("expenses_db.json", {})
    lst = list(expenses_db.get(company, []))

    # 삭제 대상 먼저 찾기
    target = None
    for e in lst:
        if str(e.get("id")) == str(eid):
            target = e
            break

    # 자동 외주 지출이면 원본 작업도 제거
    if target and str(target.get("source","")).startswith("auto_outsrc_"):
        _delete_outsourcing_job_by_auto_item(company, target)

    # 지출 행 삭제
    lst = [e for e in lst if str(e.get("id")) != str(eid)]
    expenses_db[company] = lst
    save_json("expenses_db.json", expenses_db)

    # 🔁 쿼리스트링 유지해서 리다이렉트 (페이지/검색 유지)
    params = request.args.to_dict(flat=True)
    params["tab"] = "expense_list"
    if "page" not in params:
        params["page"] = "1"
    return redirect(url_for('finance_dashboard', **params))

from urllib.parse import quote

@app.route("/finance/expense_export_xlsx", methods=["POST"])
@perm_required('manage_jobs')
def finance_expense_export_xlsx():
    from io import BytesIO
    from urllib.parse import quote
    import datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    # 열 너비 자동조절
    def _autofit_worksheet(ws, min_widths=None, max_width=60):
        min_widths = min_widths or {}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                s = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, max((len(x) for x in s.splitlines()), default=0))
            ws.column_dimensions[letter].width = max(min_widths.get(col_idx, 10),
                                                     min(max_width, max_len + 2))

    company = (session.get("company") or (get_current_user() or {}).get("company", "")).strip()
    start_s = (request.form.get("start") or "").strip()
    end_s   = (request.form.get("end") or "").strip()
    start   = _parse_date_safe(start_s) or date(date.today().year, 1, 1)
    end     = _parse_date_safe(end_s)   or date.today()

    exp_cat  = (request.form.get("exp_cat")  or "").strip().lower()
    exp_desc = (request.form.get("exp_desc") or "").strip().lower()

    def _is_auto_given(e):
        return (e.get("source") == "auto_outsrc_given") or (
            e.get("category") == "외주줬음" and e.get("auto_key")
        )

    exps_out = []
    for e in expenses_read(company):
        # 기간/검색 필터
        d = _parse_date_safe(e.get("date"))
        if not _in_range(d, start, end):              continue
        if exp_cat and exp_cat not in (e.get("category","").lower()): continue
        desc_text = str(e.get("desc") or e.get("memo") or e.get("detail") or "")
        if exp_desc and exp_desc not in desc_text.lower():            continue
        # 외주줬음 자동항목 제외
        if _is_auto_given(e):                          continue
        exps_out.append(e)

    # 엑셀
    wb = Workbook(); ws = wb.active; ws.title = "지출"
    headers = ["일자","카테고리","내역","금액(원)"]; ws.append(headers)
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = center; c.border = thin

    total_expense = 0
    for e in exps_out:
        amt = int(_to_number(e.get("amount") or 0))
        total_expense += amt
        ws.append([
            e.get("date",""),
            e.get("category",""),
            e.get("desc") or e.get("memo") or e.get("detail") or "",
            amt
        ])

    # 본문 서식
    if ws.max_row >= 2:
        amt_col_idx = 4
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for i, c in enumerate(r, start=1):
                c.border = thin
                if i == amt_col_idx:
                    c.number_format = "#,##0"; c.alignment = right
                else:
                    c.alignment = left

    # 합계 행
    ws.append(["총 지출","","", total_expense])
    for i, c in enumerate(ws[ws.max_row], start=1):
        c.border = thin
        c.font = Font(bold=True)
        c.alignment = right if i == 4 else right
        if i == 4: c.number_format = "#,##0"

    # 열너비 자동
    _autofit_worksheet(ws, min_widths={1:12, 2:16, 3:24, 4:14})

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    from urllib.parse import quote
    start_code = start.strftime('%Y%m%d')
    end_code   = end.strftime('%Y%m%d')
    filename   = f"지출내역 {start_code}_{end_code}.xlsx"
    q = quote(filename)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={q}; filename*=UTF-8''{q}"}
    )

# ====== (신규) 매출 요약 엑셀 추출 ======
@app.post("/finance/export/xlsx", endpoint="finance_export_xlsx")
def finance_export_xlsx():
    # ── imports
    from io import BytesIO
    from urllib.parse import quote
    import datetime as _dt
    import hashlib as _h
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter
    except Exception:
        return back_with_error(
            'finance_dashboard',
            "openpyxl이 없습니다. 'pip install openpyxl' 후 다시 시도해 주세요."
        )

    if 'username' not in session:
        return redirect(url_for('login'))

    # ── 공통 유틸
    def truthy(v): return str(v or "").lower() in ("1","true","y","yes","on")

    user = get_current_user() or {}
    company = (user.get('company') or '').strip()

    # ── 폼 파라미터(요약 탭과 동일)
    start = _parse_date_safe(request.form.get("start")) or date.today().replace(month=1, day=1)
    end   = _parse_date_safe(request.form.get("end"))   or date.today()
    half  = (request.form.get("half") or "").strip()

    by_worker  = truthy(request.form.get("by_worker"))
    by_machine = truthy(request.form.get("by_machine"))
    by_client  = truthy(request.form.get("by_client"))

    status_f = (request.form.get("status") or "").strip().lower()   # 'todo' / 'done' / ''
    pay_f    = (request.form.get("pay") or "").strip().lower()      # 'unpaid' / 'paid' / ''

    sel_worker   = (request.form.get("worker") or "").strip()
    input_worker = (request.form.get("worker_input") or "").strip()
    q_plate      = (request.form.get("plate") or "").strip()
    q_owner      = (request.form.get("owner") or "").strip()
    q_tenant     = (request.form.get("tenant") or "").strip()

    # 헤더 컬럼 선택
    cols_req = request.form.getlist("cols")
    header_map = {
        "date": "일자",
        "owner": "원수급자",
        "tenant": "임차인",
        "machine_name": "장비명",
        "machine_number": "차량번호",
        "worker": "기사",
        "amount": "금액(원)",
        "status": "납부여부",
    }
    cols = [c for c in cols_req if c in header_map] or list(header_map.keys())

    def _half_ok(d: date) -> bool:
        if not d: return False
        if half == "H12": return 1 <= d.month <= 6
        if half == "H34": return 7 <= d.month <= 12
        return True

    # ── 작업/수입/지출 로드
    jobs_all = load_json("jobs.json", {}).get(company, [])
    for i, j in enumerate(jobs_all): j["_idx"] = i

    incomes_db  = load_json("incomes.json",    {})
    expenses_db = load_json("expenses_db.json",{})
    inc_all = incomes_db.get(company, [])
    exp_all = expenses_db.get(company, [])

    # ── 외주 키 생성 & 외주 작업 상태 매핑(수입/지출 합계 계산용)
    def _out_key_for_job(j: dict, kind: str) -> str:
        base = "|".join([
            str(j.get("date","")),
            str(j.get("worker","")),
            str(j.get("machine_number","")),
            str(j.get("outsource_partner","")),
            kind.lower().strip()
        ])
        return "auto-" + _h.md5(base.encode("utf-8")).hexdigest()[:16]

    job_info_by_key = {}
    for j in jobs_all:
        ot = (j.get("outsource_type") or "").lower()
        if ot in ("received","given"):
            k = _out_key_for_job(j, ot)
            job_info_by_key[k] = {
                "status": (j.get("status") or "").strip(),            # 진행중/완료
                "payment_status": (j.get("payment_status") or "").strip(), # 완납/미납/부분
            }

    # ── 요약/내역 표 데이터 만들기 (외주 작업 제외)
    def _dt_key(j):
        d = (j.get("date") or "").strip()
        t = (j.get("time") or "").strip() or "00:00"
        try:
            return _dt.datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
        except:
            try:
                return _dt.datetime.strptime(d, "%Y-%m-%d")
            except:
                return _dt.datetime.min

    rows = []
    for j in jobs_all:
        d = _parse_date_safe(j.get("date"))
        if not (_in_range(d, start, end) and _half_ok(d)):
            continue
        # 외주(received/given) 제외
        if (j.get("outsource_type") or "none") in ("received", "given"):
            continue

        # 필터
        if by_worker:
            target = (sel_worker or input_worker).strip()
            if target and target not in ((j.get("worker") or j.get("driver") or "")):
                continue
        if by_machine and q_plate:
            plate = (j.get("machine_number") or j.get("plate") or "")
            if q_plate not in plate:
                continue
        if by_client:
            owner  = (j.get("client_primary") or j.get("client") or "")
            tenant = (j.get("client_tenant") or "")
            if q_owner and q_owner not in owner:    continue
            if q_tenant and q_tenant not in tenant: continue

        amt   = _amount_won(j)
        paid  = _paid_won(j)
        color = _color_by_payment(amt, paid, j.get("payment_status") or "")

        if pay_f == "unpaid" and color not in ("unpaid","partial"):
            continue
        if pay_f == "paid" and color != "paid":
            continue

        st = (j.get("status") or "진행중").strip()
        is_todo = (st != "완료")
        if status_f == "todo" and not is_todo:
            continue
        if status_f == "done" and is_todo:
            continue

        rows.append({
            "date_str": j.get("date") or "",
            "owner": (j.get("client_primary") or j.get("client") or ""),
            "tenant": j.get("client_tenant") or "",
            "machine_name": (j.get("machine_name") or j.get("machine_alias") or j.get("machine") or ""),
            "machine_number": (j.get("machine_number") or j.get("plate") or ""),
            "worker": (j.get("worker") or j.get("driver") or j.get("기사") or "-"),
            "amount_won": int(_amount_won(j)),
            "status": ("미납" if color=="unpaid" else ("부분납부" if color=="partial" else "완납")),
            "_sort": (_dt_key(j), j.get("_idx",-1)),
        })

    rows.sort(key=lambda x: x["_sort"], reverse=True)

    # ── 대시보드와 동일한 합계 계산
    # 총 매출(완료+완납, 외주받음 제외) & 미정산 총액
    completed, outstanding_all = [], 0
    for j in jobs_all:
        d = _parse_date_safe(j.get("date"))
        if not (_in_range(d, start, end) and _half_ok(d)):
            continue
        amt  = _amount_won(j)
        paid = _paid_won(j)
        outstanding_all += max(0, amt - min(amt, paid))
        if (j.get("status") or "진행중").strip() == "완료":
            if (j.get("outsource_type") or "none") != "received":
                completed.append((amt, paid))
    sales_total = int(sum(a for a, p in completed if a > 0 and p >= a))

    # 수입 합계(외주받음은 '완납'만 포함)
    def _count_income(i: dict) -> bool:
        if (i.get("source") or "") != "auto_outsrc_received":
            return True
        k = i.get("auto_key") or i.get("id")
        ji = job_info_by_key.get(k)
        return bool(ji and ji.get("payment_status") == "완납")

    income_total = int(sum(
        _to_number(i.get("amount"))
        for i in inc_all
        if _in_range(_parse_date_safe(i.get("date")), start, end) and _count_income(i)
    ))

    # 지출 합계(외주줬음 자동항목 제외)
    def _is_auto_given(e):
        return (e.get("source") == "auto_outsrc_given") or (
            e.get("category") == "외주줬음" and e.get("auto_key")
        )
    expense_total = int(sum(
        _to_number(e.get("amount"))
        for e in exp_all
        if _in_range(_parse_date_safe(e.get("date")), start, end) and not _is_auto_given(e)
    ))

    profit_total = int(sales_total - expense_total + income_total)

    # ── 워크시트 작성
    wb = Workbook()
    ws = wb.active
    ws.title = "매출요약"
    ws.freeze_panes = "A2"

    # 헤더
    ws.append([header_map[c] for c in cols])

    # 본문
    def row_to_list(r):
        mapping = {
            "date": r["date_str"],
            "owner": r["owner"],
            "tenant": r["tenant"],
            "machine_name": r["machine_name"],
            "machine_number": r["machine_number"],
            "worker": r["worker"],
            "amount": int(r["amount_won"]),
            "status": r["status"],
        }
        return [mapping[c] for c in cols]

    for r in rows:
        ws.append(row_to_list(r))

    # ── 스타일(헤더/본문/숫자서식/테두리)
    thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    # 헤더
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = center
        c.border = thin

    # 본문
    amt_col_idx = (cols.index("amount") + 1) if "amount" in cols else None
    if ws.max_row >= 2:
        for rr in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for i, c in enumerate(rr, start=1):
                c.border = thin
                if amt_col_idx and i == amt_col_idx:
                    c.number_format = "#,##0"
                    c.alignment = right
                else:
                    c.alignment = left

    # ── 하단 요약 박스
    ws.append([])
    ws.append(["요약", "금액(원)"])
    ws.append(["총 매출", sales_total])
    ws.append(["지출", expense_total])
    ws.append(["수입", income_total])
    ws.append(["미정산 총액", outstanding_all])
    ws.append(["수익 (총 매출 - 지출 + 수입)", profit_total])

    # 요약 영역 스타일
    sum_start = ws.max_row - 5  # "총 매출" 행 번호
    # 헤더(요약 제목)
    for c in ws[sum_start-1]:
        c.font = Font(bold=True)
        c.alignment = center
        c.border = thin
    # 데이터
    for r in ws.iter_rows(min_row=sum_start, max_row=ws.max_row, min_col=1, max_col=2):
        r[0].alignment = left
        r[1].alignment = right
        r[1].number_format = "#,##0"
        for c in r:
            c.border = thin

    # ── 열 너비 자동
    def _autofit_worksheet(ws, min_widths=None, max_width=60):
        min_widths = min_widths or {}
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                s = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, max((len(x) for x in s.splitlines()), default=0))
            ws.column_dimensions[letter].width = max(min_widths.get(col_idx, 10),
                                                     min(max_width, max_len + 2))

    # 컬럼별 최소 너비 조금 넉넉히
    # (원수급자/임차인은 길어질 수 있어 22)
    min_widths = {}
    for idx, key in enumerate(cols, start=1):
        if key in ("owner","tenant"): min_widths[idx] = 22
        elif key in ("machine_name","machine_number","worker"): min_widths[idx] = 14
        elif key == "date": min_widths[idx] = 12
        elif key == "amount": min_widths[idx] = 14
        else: min_widths[idx] = 12
    # 요약 박스용 두 열도 최소폭 설정
    col_cnt = ws.max_column
    min_widths[col_cnt-1] = max(min_widths.get(col_cnt-1, 12), 14)
    min_widths[col_cnt]   = max(min_widths.get(col_cnt,   12), 14)

    _autofit_worksheet(ws, min_widths=min_widths)

    # ── 반환
    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    from urllib.parse import quote
    start_code = start.strftime('%Y%m%d')
    end_code   = end.strftime('%Y%m%d')
    filename   = f"정산서 {start_code}_{end_code}.xlsx"
    q = quote(filename)

    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={q}; filename*=UTF-8''{q}"}
    )

# =========================================================
# 거래처 정리
# =========================================================
@app.post("/clients/prune")
@app.post("/manage_clients/prune")
@perm_required('manage_jobs')
def clients_prune():
    users_db = load_json('users.json', {})
    username = session.get('username')
    company = users_db.get(username, {}).get('company', '')

    kind = (request.form.get('kind') or 'primary').strip()
    try:
        days = int(request.form.get('days') or 30)
    except Exception:
        days = 30
    if days not in (30, 90, 180, 360):
        days = 30
    cutoff = date.today() - timedelta(days=days)

    partners = load_partners(company)
    owners   = partners.get('owners',  [])
    tenants  = partners.get('tenants', [])

    def canon(s: str) -> str:
        s = (s or '').strip().lower()
        s = re.sub(r'\s+', '', s)
        s = s.replace('㈜', '').replace('(주)', '').replace('주식회사', '')
        for w in ('크레인', '중기', '건설', '기계', '장비'):
            s = s.replace(w, '')
        s = s.replace('(', '').replace(')', '')
        return s

    jobs = load_json('jobs.json', {}).get(company, [])
    last_used_owner  = {}
    last_used_tenant = {}
    for j in jobs:
        d = _parse_date_safe(j.get('date'))
        if not d: continue
        o_raw = (j.get('client_primary') or j.get('client') or '').strip()
        t_raw = (j.get('client_tenant') or '').strip()
        if o_raw:
            co = canon(o_raw)
            last_used_owner[co] = max(last_used_owner.get(co, d), d)
        if t_raw:
            ct = canon(t_raw)
            last_used_tenant[ct] = max(last_used_tenant.get(ct, d), d)

    removed = []
    if kind == 'tenant':
        keep = []
        for name in tenants:
            lu = last_used_tenant.get(canon(name))
            if not lu or lu < cutoff:
                removed.append(name)
            else:
                keep.append(name)
        partners['tenants'] = keep
        label = '임차인'
    else:
        keep = []
        for name in owners:
            lu = last_used_owner.get(canon(name))
            if not lu or lu < cutoff:
                removed.append(name)
            else:
                keep.append(name)
        partners['owners'] = keep
        label = '원수급자'

    save_partners(company, partners.get('owners', []), partners.get('tenants', []))

    try:
        flash(f"최근 {days}일간 미사용 {label} {len(removed)}건 삭제", "success")
    except Exception:
        pass

    redir = url_for('manage_clients')
    from_arg = (request.form.get('from') or '').strip()
    if from_arg:
        redir += f"?from={from_arg}"
    return redirect(redir)

@app.post("/api/send-email")
def send_email_api():
    try:
        data = request.get_json(force=True, silent=True) or {}

        to_emails  = _to_list(data.get("to"))
        cc_emails  = _to_list(data.get("cc"))
        bcc_emails = _to_list(data.get("bcc"))

        subject   = (data.get("subject") or "").strip()
        body_text = (data.get("body") or "").strip()
        body_html = data.get("body_html") or None   # ← 없으면 None

        if not to_emails or not subject:
            return jsonify(success=False, error="to/subject가 비었습니다."), 400

        _send_email_link(to_emails, subject, body_text, cc_emails, bcc_emails, body_html)
        return jsonify(success=True)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500

@app.get("/__smtp_check")
@perm_required('manage_roles')
def __smtp_check():
    info = {
        "host": SMTP_HOST,
        "port": SMTP_PORT,
        "user": SMTP_USER,
        "from": SMTP_FROM,
        "has_pass": bool(SMTP_PASS),
    }
    ok = False
    err = None
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=5) as s:
            s.ehlo()
            if SMTP_USE_TLS:
                s.starttls()
                s.ehlo()
        ok = True
    except Exception as e:
        err = str(e)
    if err:
        info["error"] = err
    info["ok"] = ok
    return jsonify(info)    

@app.get("/docs")
def docs_home():
    return render_template("docs_home.html")

@app.route("/docs/upload", methods=["GET","POST"], endpoint="docs_upload")
def docs_upload():
    company = _company_name()
    dest = _ensure_company_docs_dir(company)

    if request.method == "POST":
        files = request.files.getlist("files")
        created = skipped = 0
        for f in files:
            if not f or not f.filename:
                continue
            if not _allowed_file(f.filename):
                skipped += 1
                continue

            safe = secure_filename(f.filename)
            stored = f"{uuid.uuid4().hex[:8]}-{safe}"
            path = dest / stored
            f.save(path)

            mime, _ = mimetypes.guess_type(stored)
            _docs_add(company, safe, stored, path.stat().st_size, mime or "application/octet-stream")
            created += 1

        if created:
            flash(f"{created}개 업로드 완료", "success")
        if skipped:
            flash(f"{skipped}개 파일은 허용되지 않는 확장자라 건너뜀", "error")
        return redirect(url_for("docs_manage"))

    return render_template("docs_upload.html")

# === 보내기 ===
from html import escape
from datetime import datetime
import os, mimetypes, uuid
from urllib.parse import quote as urlquote

from flask import render_template, request, redirect, url_for, flash, current_app
from werkzeug.utils import secure_filename

# 메일 보내는 내부 함수는 기존 _send_email_link 사용
# 파일 확장자 허용 체크 함수는 기존 _allowed_file 사용
# 회사명 함수는 기존 _company_name 사용

@app.route("/docs/send", methods=["GET", "POST"])
def docs_send():
    if request.method == "GET":
        return render_template("docs_send.html")

    # ========= 1) 값 읽기 =========
    to_raw    = (request.form.get("to") or "").strip()
    subject   = (request.form.get("subject") or "").strip()
    body_text = (request.form.get("body") or "").strip()
    files     = request.files.getlist("files")  # ← multiple 대응

    if not to_raw or not subject:
        flash("수신자(To)와 제목을 입력하세요.", "error")
        return redirect(url_for("docs_send"))

    to_emails = _parse_emails(to_raw)
    if not to_emails:
        flash("수신자(To) 이메일 형식이 올바르지 않습니다.", "error")
        return redirect(url_for("docs_send"))

    # ========= 2) 메일 구성 =========
    msg = EmailMessage()
    msg["From"]    = current_app.config.get("MAIL_FROM", "no-reply@jangbion.com")
    msg["To"]      = ", ".join(to_emails)
    msg["Subject"] = subject
    msg.set_content(body_text or "")

    # (참고) 표시용(메일 본문에 쓰고 싶으면 사용)
    company_display = " ".join(filter(None, [
        current_app.config.get("COMPANY_NAME_KO"),
        current_app.config.get("COMPANY_NAME_EN"),
    ])).strip() or "회사"
    today = datetime.now().strftime("%Y%m%d")

    attached = 0
    for f in files:
        if not f or not f.filename:
            continue

        # 원본 파일명(한글/공백 모두 허용). EmailMessage가 RFC2231로 인코딩해줌.
        original_name = f.filename

        # 안전 이름은 서버 저장 시만 필요하지만, 혹시 로그/백업용으로 유지
        safe_name = secure_filename(original_name) or original_name

        data = f.read()
        f.seek(0)
        if not data:
            continue

        ctype, _ = mimetypes.guess_type(original_name)
        if not ctype:
            ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)

        # 실제 첨부
        msg.add_attachment(
            data,
            maintype=maintype,
            subtype=subtype,
            filename=original_name  # ← 한글 파일명 OK (RFC2231)
        )
        attached += 1

    # ========= 3) 전송 =========
    host = current_app.config.get("SMTP_HOST")
    port = int(current_app.config.get("SMTP_PORT", 587))
    user = current_app.config.get("SMTP_USER")
    pwd  = current_app.config.get("SMTP_PASS")
    use_starttls = str(current_app.config.get("SMTP_STARTTLS", "true")).lower() != "false"

    try:
        if use_starttls:
            with smtplib.SMTP(host, port, timeout=25) as s:
                s.ehlo()
                s.starttls(context=ssl.create_default_context())
                if user and pwd:
                    s.login(user, pwd)
                s.send_message(msg)
        else:
            # SMTPS(SSL)로 쓰는 환경이면 여기로
            with smtplib.SMTP_SSL(host, port, timeout=25) as s:
                if user and pwd:
                    s.login(user, pwd)
                s.send_message(msg)

        flash(f"메일을 보냈습니다. 첨부 {attached}건.", "success")

    except Exception as e:
        current_app.logger.exception("메일 발송 실패")
        flash(f"메일 발송 실패: {e}", "error")

    # 결과 안내는 **docs_send 페이지**에서 보이도록 자기 자신으로 리다이렉트
    return redirect(url_for("docs_send"))

@app.route("/docs/register", methods=["GET", "POST"])
def docs_register():
    company = _company_name()
    dest = _ensure_company_docs_dir(company)
    uploaded = []

    if request.method == "POST":
        for f in request.files.getlist("files"):
            if not f or not f.filename:
                continue
            if not _allowed_file(f.filename):
                flash(f"허용되지 않는 확장자: {f.filename}", "error")
                continue
            safe = secure_filename(f.filename)
            stored = f"{uuid.uuid4().hex[:8]}-{safe}"
            path = dest / stored
            f.save(path)

            mime, _ = mimetypes.guess_type(stored)
            _docs_add(company, safe, stored, path.stat().st_size, mime or "application/octet-stream")
            uploaded.append(safe)

        if uploaded:
            flash(f"업로드 완료: {', '.join(uploaded)}", "success")
        return redirect(url_for("docs_register"))

    return render_template("docs_register.html")

# ---------- 서류 관리(목록/다운/삭제) ----------
@app.get("/docs/manage")
def docs_manage():
    company = _company_name()
    rows = []
    for r in _docs_list(company):
        rows.append({
            "id": r["id"],
            "name": r["filename"],      # 기존 템플릿 호환 필드
            "stored": r["stored"],
            "size": r["size"],
            "mtime": r["uploaded_at"],  # 기존 템플릿 호환 필드
            "mime": r.get("mime",""),
        })
    return render_template("docs_manage.html", rows=rows)

# 다운로드
@app.get("/docs/file/<path:key>")
def docs_file(key):
    company = _company_name()
    rec = _docs_find(company, key)
    base = _ensure_company_docs_dir(company)

    if rec:
        p = base / rec["stored"]
        dl_name = rec["filename"] or p.name
        mime = rec.get("mime") or mimetypes.guess_type(dl_name)[0] or "application/octet-stream"
    else:
        # 혹시나 파일명 직접 진입 시도(과거 링크 호환)
        p = base / secure_filename(key)
        if not p.exists():
            return "Not found", 404
        dl_name = p.name
        mime = mimetypes.guess_type(dl_name)[0] or "application/octet-stream"

    as_attach = (request.args.get("download", "1") != "0")
    return send_file(p, as_attachment=as_attach, download_name=dl_name, mimetype=mime)

# 삭제
@app.post("/docs/delete")
def docs_delete():
    company = _company_name()
    key = (request.form.get("id") or request.form.get("name") or "").strip()
    if not key:
        flash("삭제 대상이 없습니다.", "error")
        return redirect(url_for("docs_manage"))
    ok = _docs_delete(company, key)
    flash("삭제했습니다." if ok else "삭제 실패(대상을 찾지 못함).", "success" if ok else "error")
    return redirect(url_for("docs_manage"))

# =========================================================
# 디버그
# =========================================================
@app.route('/_peek_users')
def _peek_users():
    users = load_json('users.json', {})
    bosses = {k:v for k,v in users.items() if (v.get('role') == 'boss')}
    return Response(json.dumps(bosses, ensure_ascii=False, indent=2), mimetype='application/json; charset=utf-8')

@app.route('/_peek_companies')
def _peek_companies():
    return Response(json.dumps(load_json('companies.json', {}), ensure_ascii=False, indent=2),
                    mimetype='application/json; charset=utf-8')

@app.get("/__ping")
def __ping():
    return "ok", 200, {"Content-Type": "text/plain; charset=utf-8"}

@app.route("/__routes")
def __routes():
    from flask import Response
    body = "<br>".join(sorted(str(r) for r in app.url_map.iter_rules()))
    return Response(body, mimetype="text/html")

@app.get("/__env_check")
def env_check():
    email_from = (
        os.environ.get("SMTP_FROM")
        or os.environ.get("EMAIL_FROM")
        or os.environ.get("SMTP_USER")
    )
    gcs_url = os.environ.get("GCS_URL")

    ok_smtp = bool(SMTP_HOST and SMTP_PORT and email_from and SMTP_PASS)
    return jsonify({
        "email_from": email_from or None,
        "gcs_url": gcs_url or None,
        "smtp_ok": ok_smtp,
        "smtp": {
            "host": SMTP_HOST,
            "port": SMTP_PORT,
            "use_tls": SMTP_USE_TLS,
            "has_user": bool(SMTP_USER),
            "has_pass": bool(SMTP_PASS),
        },
    })

# =========================================================
# run
# =========================================================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
